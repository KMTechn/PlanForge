import pandas as pd
import numpy as np
import os
import sys
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, PanedWindow, VERTICAL, HORIZONTAL, Listbox, END, Menu, simpledialog, ttk
import tkinter.font as tkfont
import datetime
from datetime import timedelta
import math
import re
import logging
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkcalendar import Calendar, DateEntry
from tkinter import Toplevel
import requests
import zipfile
import subprocess
import threading
from queue import Queue, Empty
import json
from openpyxl.styles import PatternFill, Font, Border, Side

# =================================================================================================
# 📔 PlanForge Pro Business Workflow & Logic (v2 - 개선사항 반영)
# =================================================================================================
# (주석 생략)

# ===================================================================
# PyInstaller 빌드 환경을 위한 리소스 경로 설정 함수
# ===================================================================
def resource_path(relative_path):
    """ Get absolute path to resource, works for dev and for PyInstaller """
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# ===================================================================
# GitHub 자동 업데이트 설정
# ===================================================================
REPO_OWNER = "KMTechn"
REPO_NAME = "PlanForge"
CURRENT_VERSION = "v1.1.6" 

def check_for_updates(repo_owner: str, repo_name: str, current_version: str):
    logging.info("Checking for updates...")
    try:
        api_url = f"https://api.github.com/repos/{repo_owner}/{repo_name}/releases/latest"
        response = requests.get(api_url, timeout=5)
        response.raise_for_status()
        latest_release_data = response.json()
        latest_version = latest_release_data['tag_name']
        logging.info(f"Current version: {current_version}, Latest version: {latest_version}")
        clean_current = current_version.lower().lstrip('v').split('-')[0]
        clean_latest = latest_version.lower().lstrip('v').split('-')[0]
        if clean_latest > clean_current:
            for asset in latest_release_data['assets']:
                if asset['name'].endswith('.zip'):
                    return asset['browser_download_url'], latest_version
    except requests.exceptions.RequestException as e:
        logging.error(f"Update check failed: {e}")
    return None, None

def download_and_apply_update(url: str):
    try:
        logging.info(f"Downloading update from: {url}")
        temp_dir = os.environ.get("TEMP", "C:\\Temp")
        zip_path = os.path.join(temp_dir, "update.zip")
        response = requests.get(url, stream=True, timeout=120)
        response.raise_for_status()
        with open(zip_path, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192): f.write(chunk)
        logging.info("Download complete.")
        temp_update_folder = os.path.join(temp_dir, "temp_update")
        if os.path.exists(temp_update_folder):
            import shutil
            shutil.rmtree(temp_update_folder)
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_update_folder)
        os.remove(zip_path)
        logging.info(f"Extracted update to: {temp_update_folder}")
        application_path = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else os.path.dirname(os.path.abspath(__file__))
        updater_script_path = os.path.join(application_path, "updater.bat")
        extracted_content = os.listdir(temp_update_folder)
        new_program_folder_path = temp_update_folder
        if len(extracted_content) == 1 and os.path.isdir(os.path.join(temp_update_folder, extracted_content[0])):
            new_program_folder_path = os.path.join(temp_update_folder, extracted_content[0])
        with open(updater_script_path, "w", encoding='utf-8') as bat_file:
            bat_file.write(fr"""@echo off
chcp 65001 > nul
echo.
echo ==========================================================
echo  프로그램을 업데이트합니다. 이 창을 닫지 마세요.
echo ==========================================================
echo.
echo 잠시 후 프로그램이 자동으로 종료됩니다...
timeout /t 3 /nobreak > nul
taskkill /F /IM "{os.path.basename(sys.executable)}" > nul
echo.
echo 기존 파일을 새 파일로 교체합니다...
xcopy "{new_program_folder_path}" "{application_path}" /E /H /C /I /Y > nul
echo.
echo 임시 업데이트 파일을 삭제합니다...
rmdir /s /q "{temp_update_folder}"
echo.
echo ========================================
echo  업데이트 완료!
echo ========================================
echo.
echo 3초 후에 프로그램을 다시 시작합니다.
timeout /t 3 /nobreak > nul
start "" "{os.path.join(application_path, os.path.basename(sys.executable))}"
del "%~f0"
            """)
        logging.info("Updater batch file created.")
        subprocess.Popen(updater_script_path, creationflags=subprocess.CREATE_NEW_CONSOLE)
        sys.exit(0)
    except Exception as e:
        logging.error(f"Update application failed: {e}")
        root_alert = tk.Tk(); root_alert.withdraw()
        messagebox.showerror("업데이트 실패", f"업데이트 적용 중 오류가 발생했습니다.\n\n{e}\n\n프로그램을 다시 시작해주세요.", parent=root_alert)
        root_alert.destroy()

def run_updater(repo_owner: str, repo_name: str, current_version: str):
    def check_thread():
        download_url, new_version = check_for_updates(repo_owner, repo_name, current_version)
        if download_url:
            root_alert = tk.Tk(); root_alert.withdraw()
            if messagebox.askyesno("업데이트 발견", f"새로운 버전({new_version})이 발견되었습니다.\n지금 업데이트하시겠습니까? (현재: {current_version})", parent=root_alert):
                root_alert.destroy()
                download_and_apply_update(download_url)
            else:
                root_alert.destroy(); logging.info("User declined the update.")
        else: logging.info("No new updates found.")
    threading.Thread(target=check_thread, daemon=True).start()

# ===================================================================
# 프로그램 본체
# ===================================================================
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

class ConfigManager:
    def __init__(self):
        self.config_filepath = resource_path('config.json')
        self.config = self.load_config()

    def get_default_config(self):
        return {
            'PALLET_SIZE': 60,
            'LEAD_TIME_DAYS': 2,
            'PALLETS_PER_TRUCK': 36,
            'MAX_TRUCKS_PER_DAY': 2,
            'FONT_SIZE': 11,
            'APPEARANCE_MODE': 'System',
            'DELIVERY_DAYS': {str(i): 'True' if i < 5 else 'False' for i in range(7)},
            'NON_SHIPPING_DATES': [],
            'DAILY_TRUCK_OVERRIDES': {},
            'AUTO_SAVE_PATH': os.path.join(os.path.expanduser('~'), 'Downloads'),
            'DAILY_PALLET_OVERRIDES': {}
        }

    def load_config(self):
        try:
            with open(self.config_filepath, 'r', encoding='utf-8') as f:
                loaded_config = json.load(f)
                logging.info(f"{self.config_filepath}에서 설정을 로드했습니다.")

                default_config = self.get_default_config()
                for key, value in default_config.items():
                    if key not in loaded_config:
                        loaded_config[key] = value

                loaded_config['NON_SHIPPING_DATES'] = [datetime.datetime.strptime(d, '%Y-%m-%d').date() for d in loaded_config.get('NON_SHIPPING_DATES', [])]
                loaded_config['DAILY_TRUCK_OVERRIDES'] = {datetime.datetime.strptime(k, '%Y-%m-%d').date(): v for k, v in loaded_config.get('DAILY_TRUCK_OVERRIDES', {}).items()}
                loaded_config['DAILY_PALLET_OVERRIDES'] = {datetime.datetime.strptime(k, '%Y-%m-%d').date(): v for k, v in loaded_config.get('DAILY_PALLET_OVERRIDES', {}).items()}

                return loaded_config
        except (FileNotFoundError, json.JSONDecodeError) as e:
            logging.warning(f"설정 파일을 찾을 수 없거나 오류가 있어 기본값을 사용합니다: {e}")
            return self.get_default_config()

    def save_config(self):
        try:
            config_to_save = self.config.copy()

            config_to_save['NON_SHIPPING_DATES'] = [d.strftime('%Y-%m-%d') for d in config_to_save.get('NON_SHIPPING_DATES', [])]
            config_to_save['DAILY_TRUCK_OVERRIDES'] = {k.strftime('%Y-%m-%d'): v for k, v in config_to_save.get('DAILY_TRUCK_OVERRIDES', {}).items()}
            config_to_save['DAILY_PALLET_OVERRIDES'] = {k.strftime('%Y-%m-%d'): v for k, v in config_to_save.get('DAILY_PALLET_OVERRIDES', {}).items()}

            with open(self.config_filepath, 'w', encoding='utf-8') as f:
                json.dump(config_to_save, f, ensure_ascii=False, indent=4)
            logging.info(f"설정을 {self.config_filepath}에 저장했습니다.")
        except Exception as e:
            logging.error(f"설정 파일 저장 실패: {e}")
            messagebox.showwarning("저장 오류", f"설정 파일 저장 중 오류가 발생했습니다:\n{e}")

class PlanProcessor:
    def __init__(self, config):
        self.config = config
        self.aggregated_plan_df = None
        self.inventory_df = None
        self.simulated_plan_df = None
        self.current_filepath = ""
        self.date_cols = []
        self.inventory_date = None
        self.planning_start_date = None
        self.adjustments = []
        self.fixed_shipments = []
        self.fixed_shipment_reqs = []
        self.in_transit_inventory = []
        self.item_master_df = None
        self.allowed_models = []
        self.highlight_models = []
        self.unmet_demand_log = []
        self.optimized_additions = {}

    def _ensure_item_master_loaded(self):
        if self.item_master_df is None:
            logging.info("품목 정보(Item.csv)를 처음으로 로드합니다...")
            self._load_item_master()

    def _load_item_master(self):
        try:
            self.item_path = resource_path('assets/Item.csv')
            if not os.path.exists(self.item_path):
                raise FileNotFoundError(f"assets/Item.csv 파일을 찾을 수 없습니다. (경로: {self.item_path})")

            self.item_master_df = pd.read_csv(self.item_path)

            default_columns = {
                'Priority': ('range',),
                'SafetyStock': (0, 'int'),
                'Shipment_Unit': ('PIECE', 'str_upper'),
                'Demand_Multiplier': (1.0, 'float'),
                'Min_Shipment_Qty': (0, 'int')
            }

            for col, (default_val, *options) in default_columns.items():
                if col not in self.item_master_df.columns:
                    if default_val == 'range':
                        self.item_master_df[col] = range(1, len(self.item_master_df) + 1)
                    else:
                        self.item_master_df[col] = default_val
                else:
                    if options:
                        col_type = options[0]
                        if col_type == 'int':
                            self.item_master_df[col] = pd.to_numeric(self.item_master_df[col], errors='coerce').fillna(default_val).astype(int)
                        elif col_type == 'float':
                            self.item_master_df[col] = pd.to_numeric(self.item_master_df[col], errors='coerce').fillna(default_val)
                        elif col_type == 'str_upper':
                            self.item_master_df[col] = self.item_master_df[col].fillna(default_val).str.upper()

            if 'Spec' in self.item_master_df.columns:
                self.item_master_df['is_hmc'] = self.item_master_df['Spec'].str.contains('HMC', na=False)
                self.item_master_df.sort_values(by=['is_hmc', 'Priority'], ascending=[False, True], inplace=True)
                self.item_master_df.drop(columns=['is_hmc'], inplace=True)
            else:
                self.item_master_df.sort_values(by='Priority', inplace=True)

            self.allowed_models = self.item_master_df['Item Code'].tolist()
            if 'Spec' in self.item_master_df.columns:
                 self.highlight_models = self.item_master_df[self.item_master_df['Spec'].str.contains('HMC', na=False)]['Item Code'].tolist()

            self.item_master_df.set_index('Item Code', inplace=True)
            logging.info(f"Item.csv 로드 성공. 허용된 모델 수: {len(self.allowed_models)}. 확장 컬럼 적용 완료.")

        except Exception as e:
            messagebox.showerror("품목 정보 로드 실패", f"Item.csv 파일 처리 중 오류가 발생했습니다: {e}")
            logging.critical(f"Item.csv 로드 실패: {e}")
            raise

    def save_item_master(self):
        self._ensure_item_master_loaded()
        try:
            df_to_save = self.item_master_df.reset_index()
            if 'Priority' in df_to_save.columns:
                df_to_save.sort_values(by='Priority', inplace=True)
            df_to_save.to_csv(self.item_path, index=False, encoding='utf-8-sig')
            logging.info(f"품목 정보를 {self.item_path}에 저장했습니다.")
        except Exception as e:
            messagebox.showerror("품목 정보 저장 실패", f"Item.csv 파일 저장 중 오류 발생: {e}")
            logging.error(f"Item.csv 저장 실패: {e}")

    def process_plan_file(self):
        self._ensure_item_master_loaded()

        logging.info(f"파일 로드 시도: {self.current_filepath}")
        try:
            df_raw = pd.read_excel(self.current_filepath, sheet_name='《HCO&DIS》', header=None, engine='openpyxl')
            logging.info("원시 데이터 로드 성공. 헤더 행 탐색...")

            header_series = df_raw[11].astype(str)
            found = header_series.str.lower().str.contains('cover glass assy', na=False)
            if not found.any():
                raise ValueError("헤더 'Cover glass Assy'를 찾을 수 없습니다.")
            header_row_index = found.idxmax()

            logging.info(f"헤더 행 발견: {header_row_index}")
            df = df_raw.iloc[header_row_index:].copy()
            df.columns = df.iloc[0]
            df = df.iloc[1:].rename(columns={df.columns[11]: 'Model'})

            self.date_cols = sorted([col for col in df.columns if isinstance(col, (datetime.datetime, pd.Timestamp))])
            if not self.date_cols:
                raise ValueError("파일에서 유효한 날짜 컬럼을 찾을 수 없습니다.")

            logging.info(f"유효한 날짜 컬럼 {len(self.date_cols)}개 발견. 모델 필터링 시작...")
            df_filtered = df[df['Model'].isin(self.allowed_models)].copy()

            df_filtered.loc[:, self.date_cols] = df_filtered.loc[:, self.date_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
            agg_df = df_filtered.groupby('Model')[self.date_cols].sum()

            sorted_models = self.item_master_df.index
            reindexed_df = agg_df.reindex(sorted_models).fillna(0).astype(int)

            self.aggregated_plan_df = reindexed_df.copy()
            logging.info(f"최종 집계된 DataFrame 생성 (shape: {self.aggregated_plan_df.shape})")
            return True
        except Exception as e:
            logging.error(f"Plan file processing error: {e}")
            raise

    def load_inventory_from_text(self, text_data):
        self._ensure_item_master_loaded()
        logging.info("재고 데이터 텍스트 파싱 시작...")
        data = []
        lines = [line.strip() for line in text_data.strip().split('\n') if line.strip()]
        inventory_date = None

        for line in lines:
            date_match = re.search(r'(\d{1,2})/(\d{1,2})', line)
            if date_match:
                month = int(date_match.group(1))
                day = int(date_match.group(2))
                year = datetime.date.today().year
                inventory_date = datetime.date(year, month, day)
                break

        for line in lines:
            matches = re.findall(r'(AAA\d+)\s+.*?(\d{1,3}(?:,\d{3})*)', line)
            for match in matches:
                model, inventory_str = match
                inventory = int(inventory_str.replace(',', ''))
                data.append({'Model': model, 'Inventory': inventory})

        i = 0
        while i < len(lines) - 1:
            current_line = lines[i]
            next_line = lines[i+1]
            if current_line.startswith('AAA') and next_line.replace(',', '').isdigit():
                model = current_line
                inventory = int(next_line.replace(',', ''))
                if not any(d['Model'] == model for d in data):
                    data.append({'Model': model, 'Inventory': inventory})
                i += 2
            else:
                i += 1

        if not data:
            raise ValueError("유효한 재고 데이터를 찾을 수 없습니다. 형식을 확인하세요.")

        inventory_df_raw = pd.DataFrame(data).set_index('Model')
        self.inventory_df = inventory_df_raw[inventory_df_raw.index.isin(self.allowed_models)]
        self.inventory_date = inventory_date if inventory_date else datetime.date.today()
        logging.info(f"재고 데이터 파싱 완료. 모델 수: {len(self.inventory_df)}, 기준일: {self.inventory_date}")

    def load_inventory_from_file(self, file_path):
        self._ensure_item_master_loaded()
        logging.info(f"파일에서 재고 데이터 로드 시작: {file_path}")
        try:
            if file_path.lower().endswith('.csv'):
                df = pd.read_csv(file_path, header=None)
            elif file_path.lower().endswith(('.xlsx', '.xls')):
                df = pd.read_excel(file_path, header=None)
            else:
                raise ValueError("지원하지 않는 파일 형식입니다. (CSV, XLSX, XLS)")

            if df.shape[1] < 2:
                raise ValueError("파일은 최소 2개의 열(모델, 수량)을 포함해야 합니다.")

            df.rename(columns={0: 'Model', 1: 'Inventory'}, inplace=True)
            df['Model'] = df['Model'].astype(str)
            df['Inventory'] = pd.to_numeric(df['Inventory'], errors='coerce').fillna(0)

            inventory_df_raw = df[df['Model'].str.startswith('AAA', na=False)]
            inventory_df_raw = inventory_df_raw[['Model', 'Inventory']].set_index('Model')

            self.inventory_df = inventory_df_raw[inventory_df_raw.index.isin(self.allowed_models)]
            self.inventory_date = datetime.date.today()
            logging.info(f"파일 재고 로드 완료. 모델 수: {len(self.inventory_df)}, 기준일: {self.inventory_date}")
        except Exception as e:
            logging.error(f"Inventory file loading error: {e}")
            raise

    def run_simulation(self, adjustments=None, fixed_shipments=None, fixed_shipment_reqs=None, in_transit_inventory=None):
        self._ensure_item_master_loaded()
        logging.info("팔레트 슬롯 기반의 엄격한 적재 로직으로 시뮬레이션을 시작합니다...")
        self.adjustments = adjustments if adjustments else []
        self.fixed_shipments = fixed_shipments if fixed_shipments else []
        self.fixed_shipment_reqs = fixed_shipment_reqs if fixed_shipment_reqs else []
        self.in_transit_inventory = in_transit_inventory if in_transit_inventory else []
        self.unmet_demand_log = []
        self.optimized_additions = {} # 최적화 기록 초기화

        if self.aggregated_plan_df is None: return

        plan_df = self.aggregated_plan_df.copy()
        if self.inventory_df is not None:
            plan_df = plan_df.join(self.inventory_df, how='left').fillna({'Inventory': 0})
        else:
            plan_df = plan_df.assign(Inventory=0)
        plan_df['Inventory'] = plan_df['Inventory'].astype(int)

        sim_start_date = self.planning_start_date or (self.inventory_date if self.inventory_date else self.date_cols[0].date())
        simulation_dates = [d for d in self.date_cols if d.date() >= sim_start_date]
        logging.info(f"시뮬레이션 시작일: {sim_start_date}, 총 {len(simulation_dates)}일")

        if self.inventory_date and self.inventory_date >= sim_start_date:
            inventory_date_col = next((col for col in self.date_cols if col.date() == self.inventory_date), None)
            if inventory_date_col and inventory_date_col in plan_df.columns:
                plan_df[inventory_date_col] = 0

        if not simulation_dates: raise ValueError("시뮬레이션할 유효한 날짜가 없습니다.")

        lead_time = self.config.get('LEAD_TIME_DAYS', 2)
        pallet_size = self.config.get('PALLET_SIZE', 60)
        safety_stock = self.item_master_df['SafetyStock']

        for adj in self.adjustments:
            adj_date_dt = pd.to_datetime(adj['date'])
            if adj['model'] in plan_df.index and adj_date_dt in plan_df.columns:
                if adj['type'] == '수요': plan_df.loc[adj['model'], adj_date_dt] += adj['qty']
                elif adj['type'] == '재고': plan_df.loc[adj['model'], 'Inventory'] += adj['qty']

        demand_df = plan_df[simulation_dates].copy()
        multipliers = self.item_master_df['Demand_Multiplier']
        demand_df = demand_df.multiply(multipliers, axis='index').astype(int)

        # --- 선제적 수요 조절 로직 (역방향 패스) ---
        leveled_demand_df = demand_df.copy()
        pallet_size = self.config.get('PALLET_SIZE', 60)

        if len(simulation_dates) > 1:
            for i in range(len(simulation_dates) - 2, -1, -1):
                current_date = simulation_dates[i]
                next_date = simulation_dates[i+1]

                is_shipping_day_next = self.config['DELIVERY_DAYS'].get(str(next_date.weekday()), 'False') == 'True'
                is_non_shipping_date_next = next_date.date() in self.config['NON_SHIPPING_DATES']

                if not is_shipping_day_next or is_non_shipping_date_next:
                    leveled_demand_df[current_date] += leveled_demand_df[next_date]
                    leveled_demand_df[next_date] = 0
                    continue

                daily_max_trucks_next = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(next_date.date(), self.config.get('MAX_TRUCKS_PER_DAY', 2))
                default_pallets_per_truck_next = self.config.get('PALLETS_PER_TRUCK', 36)
                pallets_per_truck_next = self.config.get('DAILY_PALLET_OVERRIDES', {}).get(next_date.date(), default_pallets_per_truck_next)
                capacity_pallets_next_day = daily_max_trucks_next * pallets_per_truck_next

                required_pallets_next_day = np.ceil(leveled_demand_df[next_date] / pallet_size).where(leveled_demand_df[next_date] > 0, 0)

                if required_pallets_next_day.sum() > capacity_pallets_next_day:
                    shortfall_pallets = required_pallets_next_day.sum() - capacity_pallets_next_day

                    items_to_pull = required_pallets_next_day.sort_values(ascending=False).index
                    pulled_pallets_count = 0

                    for model in items_to_pull:
                        if pulled_pallets_count >= shortfall_pallets:
                            break

                        pallets_for_model = required_pallets_next_day.loc[model]
                        if pallets_for_model <= 0: continue

                        pallets_to_pull = min(pallets_for_model, shortfall_pallets - pulled_pallets_count)

                        if pallets_to_pull > 0:
                            qty_to_pull = int(pallets_to_pull * pallet_size)

                            leveled_demand_df.loc[model, current_date] += qty_to_pull
                            leveled_demand_df.loc[model, next_date] -= qty_to_pull

                            pulled_pallets_count += pallets_to_pull

                    leveled_demand_df[next_date] = leveled_demand_df[next_date].clip(lower=0)
        # --- 로직 종료 ---

        rolling_demand = leveled_demand_df.T.rolling(window=lead_time + 1, min_periods=1).sum().T
        
        inventory_over_time = pd.DataFrame(index=plan_df.index, columns=simulation_dates, dtype=np.int64)
        shipments_by_truck = {}
        customer_inventory = plan_df['Inventory'].copy().astype(np.int64)

        for date in simulation_dates:
            arriving_today = pd.Series(0, index=plan_df.index, dtype=np.int64)
            for shipment in self.in_transit_inventory:
                if shipment['arrival_date'] == date.date() and shipment['model'] in arriving_today.index:
                    arriving_today.loc[shipment['model']] += shipment['qty']
            customer_inventory += arriving_today

            daily_max_trucks = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(date.date(), self.config.get('MAX_TRUCKS_PER_DAY', 2))
            default_pallets_per_truck = self.config.get('PALLETS_PER_TRUCK', 36)
            pallets_per_truck = self.config.get('DAILY_PALLET_OVERRIDES', {}).get(date.date(), default_pallets_per_truck)
            
            is_shipping_day = self.config['DELIVERY_DAYS'].get(str(date.weekday()), 'False') == 'True'
            is_non_shipping_date = date.date() in self.config['NON_SHIPPING_DATES']
            
            total_shipments_today = pd.Series(0, index=plan_df.index, dtype=np.int64)

            if is_shipping_day and not is_non_shipping_date and daily_max_trucks > 0:
                required_for_lead_time = (rolling_demand[date] - customer_inventory + safety_stock).clip(lower=0)
                fixed_reqs_today = pd.Series(0, index=plan_df.index, dtype=np.int64)
                for req in [r for r in self.fixed_shipment_reqs if r['date'] == date.date() and r['model'] in fixed_reqs_today.index]:
                    fixed_reqs_today.loc[req['model']] += req['qty']
                
                must_ship_demand = pd.concat([required_for_lead_time, fixed_reqs_today], axis=1).max(axis=1).astype(np.int64)

                all_future_dates = [d for d in simulation_dates if d > date]
                pull_forward_demand = leveled_demand_df[all_future_dates].sum(axis=1).clip(lower=0) if all_future_dates else pd.Series(0, index=plan_df.index)

                priority_models = self.item_master_df.sort_values('Priority').index
                
                for truck_num in range(1, daily_max_trucks + 1):
                    if truck_num not in shipments_by_truck:
                        shipments_by_truck[truck_num] = pd.DataFrame(0, index=plan_df.index, columns=simulation_dates, dtype=np.int64)
                    
                    remaining_pallet_slots = pallets_per_truck

                    # STEP 1: 필수 물량 (Full Pallet)
                    for model in priority_models:
                        if remaining_pallet_slots <= 0: break
                        needed_qty = must_ship_demand.get(model, 0)
                        if needed_qty >= pallet_size:
                            num_pallets_to_load = min(math.floor(needed_qty / pallet_size), remaining_pallet_slots)
                            if num_pallets_to_load > 0:
                                qty_to_ship = num_pallets_to_load * pallet_size
                                shipments_by_truck[truck_num].loc[model, date] += qty_to_ship
                                total_shipments_today.loc[model] += qty_to_ship
                                must_ship_demand.loc[model] -= qty_to_ship
                                remaining_pallet_slots -= num_pallets_to_load

                    # STEP 2: 필수 물량 (자투리)
                    for model in priority_models:
                        if remaining_pallet_slots <= 0: break
                        if 0 < must_ship_demand.get(model, 0) < pallet_size:
                            qty_to_ship = must_ship_demand.loc[model]
                            shipments_by_truck[truck_num].loc[model, date] += qty_to_ship
                            total_shipments_today.loc[model] += qty_to_ship
                            must_ship_demand.loc[model] = 0
                            remaining_pallet_slots -= 1

                    # STEP 3: 미래 물량 (Full Pallet)
                    for model in priority_models:
                        if remaining_pallet_slots <= 0: break
                        pull_qty = pull_forward_demand.get(model, 0)
                        if pull_qty >= pallet_size:
                            num_pallets_to_load = min(math.floor(pull_qty / pallet_size), remaining_pallet_slots)
                            if num_pallets_to_load > 0:
                                qty_to_ship = num_pallets_to_load * pallet_size
                                shipments_by_truck[truck_num].loc[model, date] += qty_to_ship
                                total_shipments_today.loc[model] += qty_to_ship
                                pull_forward_demand.loc[model] -= qty_to_ship
                                remaining_pallet_slots -= num_pallets_to_load
                    
                    # STEP 4: 기존 자투리 팔레트 효율화 (Topping Off) - 미래 재고 없어도 임의로 채움
                    items_on_this_truck = shipments_by_truck[truck_num][date]
                    partial_pallet_items = items_on_this_truck[(items_on_this_truck > 0) & (items_on_this_truck % pallet_size != 0)]
                    
                    for model, current_qty in partial_pallet_items.items():
                        space_on_pallet = pallet_size - (current_qty % pallet_size)
                        
                        if space_on_pallet > 0:
                            qty_to_add = space_on_pallet
                            shipments_by_truck[truck_num].loc[model, date] += qty_to_add
                            total_shipments_today.loc[model] += qty_to_add
                            
                            key = (date.date(), truck_num, model)
                            self.optimized_additions[key] = self.optimized_additions.get(key, 0) + qty_to_add

                unmet_demand = must_ship_demand.clip(lower=0)
                if unmet_demand.sum() > 0:
                    for model, qty in unmet_demand[unmet_demand > 0].items():
                        log_entry = {'date': date.date(), 'model': model, 'unmet_qty': int(qty)}
                        self.unmet_demand_log.append(log_entry)
            
            daily_demand = demand_df[date]
            customer_inventory += total_shipments_today - daily_demand
            inventory_over_time[date] = customer_inventory

        result_df = plan_df.copy()
        for date in simulation_dates:
            date_str = date.strftime("%m%d")
            result_df[f'재고_{date_str}'] = inventory_over_time[date]
            max_trucks = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(date.date(), self.config.get('MAX_TRUCKS_PER_DAY', 2))
            for truck_num in range(1, max_trucks + 1):
                col_name = f'출고_{truck_num}차_{date_str}'
                if truck_num in shipments_by_truck and date in shipments_by_truck[truck_num].columns:
                    result_df[col_name] = shipments_by_truck[truck_num][date]
                else:
                    result_df[col_name] = 0

        self.simulated_plan_df = result_df.fillna(0).astype(int)
        logging.info("시뮬레이션 완료.")
        
    def find_stabilization_proposal(self, max_truck_limit=3):
        self._ensure_item_master_loaded()
        if self.simulated_plan_df is None: return None

        if self.unmet_demand_log:
            first_failure = self.unmet_demand_log[0]
            shipping_date = first_failure['date']
            current_max = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(shipping_date, self.config.get('MAX_TRUCKS_PER_DAY'))

            if current_max >= max_truck_limit:
                return {"type": "error", "message": f"안정화 실패: {shipping_date}의 차수가 이미 최대({max_truck_limit}회)입니다."}

            return {
                "type": "proposal",
                "reason": f"{shipping_date.strftime('%m/%d')}의 필수 출고량 부족",
                "date": shipping_date,
                "new_truck_count": current_max + 1
            }

        has_shortage, fix_info = self.find_and_propose_fix(max_truck_limit)
        if has_shortage:
            if 'error' in fix_info:
                return {"type": "error", "message": f"안정화 실패: {fix_info['error']}"}

            shipping_date = fix_info['shipping_date']
            current_max = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(shipping_date, self.config.get('MAX_TRUCKS_PER_DAY'))

            return {
                "type": "proposal",
                "reason": f"{fix_info['shortage_date'].strftime('%m/%d')}의 '{fix_info['model']}' 재고 부족",
                "date": shipping_date,
                "new_truck_count": current_max + 1
            }

        return {"type": "stable", "message": "현재 계획은 안정적입니다."}

    def find_and_propose_fix(self, max_truck_limit=3):
        self._ensure_item_master_loaded()
        if self.simulated_plan_df is None:
            return False, None

        df = self.simulated_plan_df
        inventory_cols = sorted([c for c in df.columns if isinstance(c, str) and c.startswith('재고_')])
        first_shortage_info = None

        for inv_col in inventory_cols:
            date_str = inv_col.split('_')[1]
            year = self.date_cols[0].year
            try:
                current_date = datetime.datetime.strptime(f"{year}-{date_str}", "%Y-%m%d").date()
            except ValueError:
                current_date = datetime.datetime.strptime(f"{year+1}-{date_str}", "%Y-%m%d").date()

            sorted_models = self.item_master_df.index
            models_in_df = self.item_master_df.index.intersection(df.index)
            shortage_series = df.loc[models_in_df, inv_col] < self.item_master_df.loc[models_in_df, 'SafetyStock']

            if shortage_series.any():
                shortage_models = shortage_series[shortage_series].index
                model_to_fix = next((model for model in sorted_models if model in shortage_models), None)

                if model_to_fix:
                    first_shortage_info = {"model": model_to_fix, "shortage_date": current_date}
                    break

        if not first_shortage_info:
            return False, None

        shortage_date = first_shortage_info['shortage_date']

        sim_start_date = self.planning_start_date or (self.inventory_date if self.inventory_date else self.date_cols[0].date())
        candidate_days = []
        check_date = shortage_date - datetime.timedelta(days=1)

        while check_date >= sim_start_date:
            is_shipping_day = self.config.get('DELIVERY_DAYS', {}).get(str(check_date.weekday()), 'False') == 'True'
            is_non_shipping_date = check_date in self.config.get('NON_SHIPPING_DATES', [])
            if is_shipping_day and not is_non_shipping_date:
                current_trucks = self.config.get('DAILY_TRUCK_OVERRIDES', {}).get(check_date, self.config.get('MAX_TRUCKS_PER_DAY'))
                candidate_days.append({"date": check_date, "trucks": current_trucks})
            check_date -= datetime.timedelta(days=1)

        if not candidate_days:
            return False, {"error": f"{first_shortage_info['model']} 부족({shortage_date})을 해결할 이전 납품일 없음"}

        eligible_candidates = [day for day in candidate_days if day['trucks'] < max_truck_limit]

        if not eligible_candidates:
            return False, {"error": f"재고 부족을 해결할 수 없습니다. 모든 유효한 이전 납품일의 차수가 최대({max_truck_limit}회)입니다."}

        eligible_candidates.sort(key=lambda x: (x['trucks'], -x['date'].toordinal()))

        best_candidate = eligible_candidates[0]

        fix_details = {
            "model": first_shortage_info['model'],
            "shortage_date": shortage_date,
            "shipping_date": best_candidate['date']
        }

        return True, fix_details

class SearchableComboBox(ctk.CTkFrame):
    def __init__(self, parent, values, font=None):
        super().__init__(parent, fg_color="transparent")
        self.values = sorted(values)
        self.current_value = ""

        self.entry = ctk.CTkEntry(self, placeholder_text="모델 검색 또는 선택...", font=font)
        self.entry.pack(fill="x")
        self.entry.bind("<KeyRelease>", self.on_key_release)

        self.listbox = Listbox(self, height=5, font=(font.cget("family"), font.cget("size")) if font else None)
        self.listbox.bind("<<ListboxSelect>>", self.on_listbox_select)

        self.entry.bind("<FocusOut>", self.hide_listbox)
        self.listbox.bind("<FocusOut>", self.hide_listbox)
        self.bind("<FocusOut>", self.hide_listbox)

    def on_key_release(self, event=None):
        search_term = self.entry.get().lower()
        if not search_term:
            self.listbox.pack_forget()
            return

        matching_values = [v for v in self.values if search_term in v.lower()]

        if matching_values:
            self.listbox.delete(0, END)
            for val in matching_values:
                self.listbox.insert(END, val)
            self.listbox.pack(fill="x", expand=True)
            self.listbox.lift()
        else:
            self.listbox.pack_forget()

    def on_listbox_select(self, event=None):
        if self.listbox.curselection():
            self.current_value = self.listbox.get(self.listbox.curselection())
            self.entry.delete(0, END)
            self.entry.insert(0, self.current_value)
            self.listbox.pack_forget()
            self.focus()

    def hide_listbox(self, event=None):
        self.after(200, lambda: self.listbox.pack_forget() if self.focus_get() != self.listbox else None)

    def get(self):
        return self.entry.get()

class LastShipmentDialog(ctk.CTkToplevel):
    def __init__(self, parent, available_dates, font_normal=None):
        super().__init__(parent)
        self.title("납품 완료 기준일 선택")
        self.geometry("400x180")
        self.result = None

        self.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(self, text="이미 납품이 완료된 마지막 날짜를 선택해주세요.\n선택한 날짜의 다음날부터 계획을 시작합니다.", justify="left", font=font_normal).pack(padx=20, pady=10)

        formatted_dates = ["(선택 안함 - 전체 계획 생성)"] + [d.strftime('%Y-%m-%d') for d in available_dates]
        self.date_combo = ctk.CTkComboBox(self, values=formatted_dates, width=360, font=font_normal)
        self.date_combo.pack(padx=20, pady=10, fill="x")
        self.date_combo.set(formatted_dates[0])

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.pack(padx=20, pady=10, anchor="e")
        ctk.CTkButton(button_frame, text="확인", command=self.ok_event, font=font_normal).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="취소", command=self.cancel_event, fg_color="gray", font=font_normal).pack(side="left")

        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self.cancel_event)

    def ok_event(self):
        selected = self.date_combo.get()
        if "(선택 안함" in selected:
            self.result = "None"
        else:
            try:
                self.result = datetime.datetime.strptime(selected, '%Y-%m-%d').date()
            except (ValueError, TypeError):
                self.result = None
        self.destroy()

    def cancel_event(self):
        self.result = None
        self.destroy()

class AdjustmentDialog(ctk.CTkToplevel):
    def __init__(self, parent, models, font_normal=None):
        super().__init__(parent)
        self.models = models
        self.adjustments = []
        self.result = None
        self.title("수동 조정 입력")
        self.geometry("600x500")
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=0, column=0, columnspan=2, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="모델:", font=font_normal).grid(row=0, column=0, padx=5, pady=5)
        self.model_combo = SearchableComboBox(input_frame, values=self.models, font=font_normal)
        self.model_combo.grid(row=0, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(input_frame, text="날짜 (YYYY-MM-DD):", font=font_normal).grid(row=1, column=0, padx=5, pady=5)
        self.date_entry = ctk.CTkEntry(input_frame, placeholder_text=datetime.date.today().strftime('%Y-%m-%d'), font=font_normal)
        self.date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(input_frame, text="수량:", font=font_normal).grid(row=2, column=0, padx=5, pady=5)
        self.qty_entry = ctk.CTkEntry(input_frame, font=font_normal)
        self.qty_entry.grid(row=2, column=1, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(input_frame, text="타입:", font=font_normal).grid(row=3, column=0, padx=5, pady=5)
        self.type_combo = ctk.CTkComboBox(input_frame, values=['재고', '수요', '고정 출고'], font=font_normal)
        self.type_combo.grid(row=3, column=1, padx=5, pady=5, sticky="ew")

        button_frame = ctk.CTkFrame(self)
        button_frame.grid(row=1, column=0, columnspan=2, padx=10, pady=5)
        ctk.CTkButton(button_frame, text="추가", command=self.add_adjustment, font=font_normal).pack()

        self.listbox = Listbox(self, height=10, font=(font_normal.cget("family"), font_normal.cget("size")) if font_normal else None)
        self.listbox.grid(row=2, column=0, columnspan=2, padx=10, pady=5, sticky="nsew")

        ok_cancel_frame = ctk.CTkFrame(self, fg_color="transparent")
        ok_cancel_frame.grid(row=3, column=0, columnspan=2, padx=10, pady=10, sticky="e")
        ctk.CTkButton(ok_cancel_frame, text="확인", command=self.ok_event, font=font_normal).pack(side="left", padx=10)
        ctk.CTkButton(ok_cancel_frame, text="취소", command=self.cancel_event, fg_color="gray", font=font_normal).pack(side="left")

        self.transient(parent)
        self.grab_set()

    def add_adjustment(self):
        model = self.model_combo.get()
        date_str = self.date_entry.get() or self.date_entry.cget("placeholder_text")
        qty_str = self.qty_entry.get()
        adj_type = self.type_combo.get()
        if not all([model, date_str, qty_str, adj_type]):
            messagebox.showwarning("입력 오류", "모든 필드를 채워주세요.", parent=self)
            return
        try:
            adj_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            quantity = int(qty_str)
        except ValueError:
            messagebox.showwarning("형식 오류", "날짜는 'YYYY-MM-DD', 수량은 숫자로 입력해야 합니다.", parent=self)
            return
        adj = {'model': model, 'date': adj_date, 'qty': quantity, 'type': adj_type}
        self.adjustments.append(adj)
        self.listbox.insert(END, f"{adj['type']} | {adj['date']}, {adj['model']}, {adj['qty']:,}")
        self.qty_entry.delete(0, END)
        logging.info(f"조정 항목 추가: {adj}")

    def ok_event(self):
        self.result = self.adjustments
        self.destroy()

    def cancel_event(self):
        self.result = None
        self.destroy()

class InventoryInputDialog(ctk.CTkToplevel):
    def __init__(self, parent, font_normal=None):
        super().__init__(parent)
        self.title("재고 데이터 입력")
        self.geometry("450x350")
        self.result = None
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        prompt_frame = ctk.CTkFrame(self, fg_color="transparent")
        prompt_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(prompt_frame, text="재고 데이터를 붙여넣거나 파일을 불러오세요.", font=font_normal).pack(side="left")
        ctk.CTkButton(prompt_frame, text="파일에서 불러오기", command=self.load_file, font=font_normal).pack(side="right")

        self.textbox = ctk.CTkTextbox(self, width=430, height=200, font=font_normal)
        self.textbox.grid(row=1, column=0, padx=10, pady=(0, 10), sticky="nsew")

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=2, column=0, padx=10, pady=(0, 10), sticky="e")
        ctk.CTkButton(button_frame, text="확인", command=self.ok_event, font=font_normal).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="취소", command=self.cancel_event, fg_color="gray", font=font_normal).pack(side="left", padx=5)

        self.transient(parent)
        self.grab_set()
        self.textbox.focus()

    def load_file(self):
        file_path = filedialog.askopenfilename(
            title="재고 파일 선택 (Excel, CSV)",
            filetypes=(("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*"))
        )
        if file_path:
            self.result = ('file', file_path)
            self.destroy()

    def ok_event(self):
        pasted_text = self.textbox.get("1.0", "end-1c")
        if pasted_text:
            self.result = ('text', pasted_text)
            self.destroy()
        else:
            messagebox.showwarning("입력 오류", "데이터를 입력하거나 파일을 선택해주세요.", parent=self)

    def cancel_event(self):
        self.result = None
        self.destroy()

class InTransitDialog(ctk.CTkToplevel):
    def __init__(self, parent, models, lead_time, inventory_date, font_normal=None):
        super().__init__(parent)
        self.title("납품 예정 정보 입력 (운송 중 재고)")
        self.geometry("600x500")
        self.result = []
        self.models = models
        self.lead_time = lead_time
        self.inventory_date = inventory_date

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(2, weight=1)

        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="모델:", font=font_normal).grid(row=0, column=0, padx=5, pady=5, sticky="w")
        self.model_combo = SearchableComboBox(input_frame, values=self.models, font=font_normal)
        self.model_combo.grid(row=0, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        ctk.CTkLabel(input_frame, text="도착 예정일:", font=font_normal).grid(row=1, column=0, padx=5, pady=5, sticky="w")

        self.arrival_day_var = tk.StringVar()
        arrival_dates = [(self.inventory_date + timedelta(days=i)).strftime('%Y-%m-%d') for i in range(1, self.lead_time + 1)]
        self.arrival_date_combo = ctk.CTkComboBox(input_frame, values=arrival_dates, variable=self.arrival_day_var, font=font_normal)
        self.arrival_date_combo.grid(row=1, column=1, columnspan=2, padx=5, pady=5, sticky="ew")
        if arrival_dates:
            self.arrival_date_combo.set(arrival_dates[0])

        ctk.CTkLabel(input_frame, text="도착 예정 수량:", font=font_normal).grid(row=2, column=0, padx=5, pady=5, sticky="w")
        self.qty_entry = ctk.CTkEntry(input_frame, font=font_normal)
        self.qty_entry.grid(row=2, column=1, columnspan=2, padx=5, pady=5, sticky="ew")

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=1, column=0, padx=10, pady=5)
        ctk.CTkButton(button_frame, text="추가", command=self.add_in_transit_item, font=font_normal).pack()

        self.listbox = Listbox(self, font=(font_normal.cget("family"), font_normal.cget("size")) if font_normal else None)
        self.listbox.grid(row=2, column=0, padx=10, pady=5, sticky="nsew")

        ok_cancel_frame = ctk.CTkFrame(self, fg_color="transparent")
        ok_cancel_frame.grid(row=3, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkButton(ok_cancel_frame, text="확인", command=self.ok_event, font=font_normal).pack(side="left", padx=10)
        ctk.CTkButton(ok_cancel_frame, text="취소(건너뛰기)", command=self.cancel_event, fg_color="gray", font=font_normal).pack(side="left")

        self.transient(parent)
        self.grab_set()

    def add_in_transit_item(self):
        model = self.model_combo.get()
        arrival_date_str = self.arrival_day_var.get()
        qty_str = self.qty_entry.get()

        if not all([model, arrival_date_str, qty_str]):
            messagebox.showwarning("입력 오류", "모든 필드를 채워주세요.", parent=self)
            return
        try:
            arrival_date = datetime.datetime.strptime(arrival_date_str, '%Y-%m-%d').date()
            quantity = int(qty_str)
            if quantity <= 0: raise ValueError
        except ValueError:
            messagebox.showwarning("형식 오류", "수량은 0보다 큰 숫자로 입력해야 합니다.", parent=self)
            return

        item = {'model': model, 'arrival_date': arrival_date, 'qty': quantity}
        self.result.append(item)
        self.listbox.insert(END, f"{item['arrival_date']} 도착 | {item['model']}, {item['qty']:,}개")
        self.qty_entry.delete(0, END)
        logging.info(f"납품 예정 정보 추가: {item}")

    def ok_event(self):
        self.destroy()

    def cancel_event(self):
        self.result = []
        self.destroy()

class HolidayDialog(ctk.CTkToplevel):
    def __init__(self, parent, non_shipping_dates, font_normal=None):
        super().__init__(parent)
        self.title("휴무일/공휴일 설정")
        self.geometry("300x350")
        self.result = None
        self.non_shipping_dates = [d for d in non_shipping_dates if isinstance(d, datetime.date)]

        self.cal = Calendar(self, selectmode='day', font=(font_normal.cget("family"), font_normal.cget("size")) if font_normal else "Arial 8")
        self.cal.pack(padx=10, pady=10, fill='x', expand=True)

        for date in self.non_shipping_dates:
            self.cal.calevent_add(date, 'holiday', 'holiday')
        self.cal.tag_config('holiday', background='red', foreground='white')

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.pack(pady=10)
        ctk.CTkButton(button_frame, text="추가/제거", command=self.toggle_date, font=font_normal).pack(side="left", padx=5)
        ctk.CTkButton(button_frame, text="확인", command=self.ok_event, font=font_normal).pack(side="left", padx=5)
        self.transient(parent)
        self.grab_set()

    def toggle_date(self):
        selected_date = self.cal.selection_get()
        if not selected_date: return

        if selected_date in self.non_shipping_dates:
            self.non_shipping_dates.remove(selected_date)
            event_ids = self.cal.get_calevents(selected_date, 'holiday')
            for event_id in event_ids:
                self.cal.calevent_remove(event_id)
        else:
            self.non_shipping_dates.append(selected_date)
            self.cal.calevent_add(selected_date, 'holiday', 'holiday')
        logging.info(f"휴무일 설정 변경: {self.non_shipping_dates}")

    def ok_event(self):
        self.result = self.non_shipping_dates
        self.destroy()

class DailyTruckDialog(ctk.CTkToplevel):
    def __init__(self, parent, truck_overrides, pallet_overrides, default_pallets, font_normal=None):
        super().__init__(parent)
        self.truck_overrides = truck_overrides.copy()
        self.pallet_overrides = pallet_overrides.copy()
        self.default_pallets = default_pallets
        self.result = None
        self.title("일자별 최대 차수 및 팔레트 수 설정")
        self.geometry("600x450")

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)
        
        font_tuple = (font_normal.cget("family"), font_normal.cget("size")) if font_normal else None

        input_frame = ctk.CTkFrame(self)
        input_frame.grid(row=0, column=0, padx=10, pady=10, sticky="ew")

        ctk.CTkLabel(input_frame, text="날짜:", font=font_normal).pack(side="left", padx=5)
        self.date_entry = DateEntry(input_frame, date_pattern='y-mm-dd', width=12, font=font_tuple)
        self.date_entry.pack(side="left", padx=5)

        ctk.CTkLabel(input_frame, text="최대 차수:", font=font_normal).pack(side="left", padx=5)
        self.truck_entry = ctk.CTkEntry(input_frame, width=50, font=font_normal)
        self.truck_entry.pack(side="left", padx=5)

        ctk.CTkLabel(input_frame, text="트럭당 PL 수:", font=font_normal).pack(side="left", padx=5)
        self.pallet_entry = ctk.CTkEntry(input_frame, width=50, placeholder_text=f"기본값({self.default_pallets})", font=font_normal)
        self.pallet_entry.pack(side="left", padx=5)

        ctk.CTkButton(input_frame, text="추가/수정", command=self.add_override, font=font_normal).pack(side="left", padx=10)

        list_frame = ctk.CTkFrame(self)
        list_frame.grid(row=1, column=0, padx=10, pady=5, sticky="nsew")
        list_frame.grid_columnconfigure(0, weight=1)
        list_frame.grid_rowconfigure(0, weight=1)

        self.listbox = Listbox(list_frame, height=15, font=font_tuple)
        self.listbox.grid(row=0, column=0, sticky="nsew")
        self.update_listbox()

        ctk.CTkButton(list_frame, text="선택 항목 삭제", command=self.remove_override, font=font_normal).grid(row=1, column=0, pady=5)

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=2, column=0, padx=10, pady=10, sticky="e")
        ctk.CTkButton(button_frame, text="저장", command=self.ok_event, font=font_normal).pack(side="left", padx=10)
        ctk.CTkButton(button_frame, text="취소", command=self.cancel_event, fg_color="gray", font=font_normal).pack(side="left")

        self.transient(parent)
        self.grab_set()

    def update_listbox(self):
        self.listbox.delete(0, END)
        all_dates = sorted(list(set(self.truck_overrides.keys()) | set(self.pallet_overrides.keys())))

        for date in all_dates:
            trucks = self.truck_overrides.get(date)
            pallets = self.pallet_overrides.get(date)

            truck_str = f"{trucks}차" if trucks is not None else "기본 차수"
            pallet_str = f"{pallets} PL" if pallets is not None else f"기본 {self.default_pallets} PL"

            self.listbox.insert(END, f"{date.strftime('%Y-%m-%d')}  ->  {truck_str} / {pallet_str}")

    def add_override(self):
        try:
            date = self.date_entry.get_date()
            trucks_str = self.truck_entry.get()
            pallets_str = self.pallet_entry.get()

            if trucks_str:
                trucks = int(trucks_str)
                if trucks < 0: raise ValueError
                self.truck_overrides[date] = trucks
            else:
                if date in self.truck_overrides:
                    del self.truck_overrides[date]

            if pallets_str:
                pallets = int(pallets_str)
                if pallets < 0: raise ValueError
                self.pallet_overrides[date] = pallets
            else:
                if date in self.pallet_overrides:
                    del self.pallet_overrides[date]

            self.update_listbox()
            self.truck_entry.delete(0, 'end')
            self.pallet_entry.delete(0, 'end')

        except (ValueError, TypeError):
            messagebox.showwarning("입력 오류", "유효한 날짜와 0 이상의 숫자를 입력하세요.", parent=self)

    def remove_override(self):
        selected_indices = self.listbox.curselection()
        if not selected_indices:
            return
        selected_text = self.listbox.get(selected_indices[0])
        date_str = selected_text.split(" ")[0]
        date_obj = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()

        if date_obj in self.truck_overrides:
            del self.truck_overrides[date_obj]
        if date_obj in self.pallet_overrides:
            del self.pallet_overrides[date_obj]

        self.update_listbox()

    def ok_event(self):
        self.result = {
            'trucks': self.truck_overrides,
            'pallets': self.pallet_overrides
        }
        self.destroy()

    def cancel_event(self):
        self.result = None
        self.destroy()

class SafetyStockDialog(ctk.CTkToplevel):
    def __init__(self, parent, item_master_df, font_normal=None, font_bold=None):
        super().__init__(parent)
        self.title("품목별 최소 재고 설정")
        self.geometry("500x600")
        self.result = None
        self.item_master_df = item_master_df.copy()
        self.entries = {}

        search_frame = ctk.CTkFrame(self)
        search_frame.pack(fill='x', padx=10, pady=5)
        ctk.CTkLabel(search_frame, text="품목 검색:", font=font_normal).pack(side='left')
        self.search_entry = ctk.CTkEntry(search_frame, font=font_normal)
        self.search_entry.pack(side='left', fill='x', expand=True, padx=5)
        self.search_entry.bind('<KeyRelease>', self.filter_items)

        header_frame = ctk.CTkFrame(self, fg_color="gray20")
        header_frame.pack(fill='x', padx=10, pady=(5,0))
        ctk.CTkLabel(header_frame, text="품목 코드", anchor='w', text_color="white", font=font_bold).pack(side='left', expand=True, fill='x', padx=5)
        ctk.CTkLabel(header_frame, text="최소 재고 수량", anchor='e', text_color="white", font=font_bold).pack(side='right', padx=20)

        self.scrollable_frame = ctk.CTkScrollableFrame(self)
        self.scrollable_frame.pack(expand=True, fill='both', padx=10, pady=(0,10))
        self.item_widgets = {}
        self.font_normal = font_normal
        self.populate_items()

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.pack(fill='x', padx=10, pady=10)
        ctk.CTkButton(button_frame, text="전체 0으로 설정", command=self.set_all_zero, fg_color="gray", font=font_normal).pack(side='left', padx=10)
        ctk.CTkButton(button_frame, text="저장", command=self.save_and_close, font=font_normal).pack(side='right', padx=10)
        ctk.CTkButton(button_frame, text="취소", command=self.cancel, fg_color="gray", font=font_normal).pack(side='right')

        self.transient(parent)
        self.grab_set()

    def populate_items(self, filter_text=""):
        for frame in self.item_widgets.values():
            frame.destroy()
        self.item_widgets.clear()
        self.entries.clear()

        df_to_show = self.item_master_df
        if filter_text:
            df_to_show = df_to_show[df_to_show.index.str.contains(filter_text, case=False)]

        for item_code, row in df_to_show.iterrows():
            frame = ctk.CTkFrame(self.scrollable_frame, fg_color="transparent")
            frame.pack(fill='x', pady=2)
            self.item_widgets[item_code] = frame

            label = ctk.CTkLabel(frame, text=item_code, anchor='w', font=self.font_normal)
            label.pack(side='left', padx=5)

            entry = ctk.CTkEntry(frame, width=100, justify='right', font=self.font_normal)
            entry.insert(0, str(row['SafetyStock']))
            entry.pack(side='right', padx=5)
            self.entries[item_code] = entry

    def filter_items(self, event=None):
        filter_text = self.search_entry.get()
        self.populate_items(filter_text)

    def set_all_zero(self):
        for entry in self.entries.values():
            entry.delete(0, END)
            entry.insert(0, "0")

    def save_and_close(self):
        try:
            for item_code, entry in self.entries.items():
                value = int(entry.get())
                self.item_master_df.loc[item_code, 'SafetyStock'] = value
            self.result = self.item_master_df
            self.destroy()
        except ValueError:
            messagebox.showerror("입력 오류", "최소 재고는 숫자로만 입력해야 합니다.", parent=self)

    def cancel(self):
        self.result = None
        self.destroy()

class ItemOrderDialog(ctk.CTkToplevel):
    def __init__(self, parent, item_list, highlight_items, font_normal=None):
        super().__init__(parent)
        self.title("품목 표시 순서 설정")
        self.geometry("400x600")
        self.result = None

        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)

        main_frame = ctk.CTkFrame(self)
        main_frame.grid(row=0, column=0, columnspan=2, sticky="nsew", padx=10, pady=10)
        main_frame.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(0, weight=1)

        listbox_font_tuple = (font_normal.cget("family"), font_normal.cget("size")) if font_normal else ("Malgun Gothic", 12)
        self.listbox = Listbox(main_frame, selectmode=tk.SINGLE, font=listbox_font_tuple)
        self.listbox.grid(row=0, column=0, sticky="nsew")

        for item in item_list:
            self.listbox.insert(END, item)

        for i in range(self.listbox.size()):
            item = self.listbox.get(i)
            if item in highlight_items:
                self.listbox.itemconfig(i, {'bg':'#D6EAF8', 'fg':'#154360'})

        button_frame = ctk.CTkFrame(self, fg_color="transparent")
        button_frame.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 10))
        button_frame.grid_columnconfigure((0,1,2,3), weight=1)

        ctk.CTkButton(button_frame, text="▲ 맨 위로", command=self.move_top, font=font_normal).grid(row=0, column=0, padx=2, pady=2)
        ctk.CTkButton(button_frame, text="△ 위로", command=self.move_up, font=font_normal).grid(row=0, column=1, padx=2, pady=2)
        ctk.CTkButton(button_frame, text="▽ 아래로", command=self.move_down, font=font_normal).grid(row=0, column=2, padx=2, pady=2)
        ctk.CTkButton(button_frame, text="▼ 맨 아래로", command=self.move_bottom, font=font_normal).grid(row=0, column=3, padx=2, pady=2)

        ok_cancel_frame = ctk.CTkFrame(self, fg_color="transparent")
        ok_cancel_frame.grid(row=2, column=0, sticky="e", padx=10, pady=10)
        ctk.CTkButton(ok_cancel_frame, text="저장", command=self.save_order, font=font_normal).pack(side="left", padx=5)
        ctk.CTkButton(ok_cancel_frame, text="취소", command=self.cancel, fg_color="gray", font=font_normal).pack(side="left")

        self.transient(parent)
        self.grab_set()

    def move_up(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx > 0:
                item = self.listbox.get(idx)
                self.listbox.delete(idx)
                self.listbox.insert(idx - 1, item)
                self.listbox.selection_set(idx - 1)
        except IndexError:
            pass

    def move_down(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx < self.listbox.size() - 1:
                item = self.listbox.get(idx)
                self.listbox.delete(idx)
                self.listbox.insert(idx + 1, item)
                self.listbox.selection_set(idx + 1)
        except IndexError:
            pass

    def move_top(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx > 0:
                item = self.listbox.get(idx)
                self.listbox.delete(idx)
                self.listbox.insert(0, item)
                self.listbox.selection_set(0)
        except IndexError:
            pass

    def move_bottom(self):
        try:
            idx = self.listbox.curselection()[0]
            if idx < self.listbox.size() - 1:
                item = self.listbox.get(idx)
                self.listbox.delete(idx)
                self.listbox.insert(END, item)
                self.listbox.selection_set(END)
        except IndexError:
            pass

    def save_order(self):
        self.result = list(self.listbox.get(0, END))
        self.destroy()

    def cancel(self):
        self.result = None
        self.destroy()

class ProductionPlannerApp(ctk.CTk):
    def __init__(self, config_manager):
        super().__init__()
        self.config_manager = config_manager
        self.processor = PlanProcessor(self.config_manager.config)

        saved_mode = self.config_manager.config.get('APPEARANCE_MODE', 'System')
        ctk.set_appearance_mode(saved_mode)

        self.current_step = 0
        self.current_file = "파일이 로드되지 않았습니다."
        self.base_font_size = self.config_manager.config.get('FONT_SIZE', 11)
        self.font_big_bold = ctk.CTkFont(size=20, weight="bold")
        self.font_normal = ctk.CTkFont(size=self.base_font_size)
        self.font_small = ctk.CTkFont(size=self.base_font_size - 1)
        self.font_bold = ctk.CTkFont(size=self.base_font_size, weight="bold")
        self.font_italic = ctk.CTkFont(size=self.base_font_size, slant="italic")
        self.font_kpi = ctk.CTkFont(size=14, weight="bold")
        self.font_header = ctk.CTkFont(size=self.base_font_size + 1, weight="bold")
        self.font_edit = ctk.CTkFont(size=self.base_font_size, weight="bold")

        self.auto_save_path_var = tk.StringVar()
        self.display_mode = 'sum' # 'sum' 또는 'adjustment'

        self.title(f"PlanForge Pro - 출고계획 시스템 ({CURRENT_VERSION})")
        self.geometry("1800x1000")

        ctk.set_default_color_theme("blue")

        self.is_task_running = False
        self.thread_queue = Queue()

        self.sidebar_visible = True
        self.inventory_text_backup = None
        self.last_selected_model = None

        self.animation_job = None
        self.animation_chars = [' .', ' ..', ' ...']
        self.animation_idx = 0
        self.status_message_base = ""

        self.stabilization_iteration = 0
        self.stabilization_active = False
        self.applied_fixes = set()

        self.calculated_sidebar_width = 320

        self.warnings_visible = False

        self.create_widgets()
        self.update_status_bar()
        self.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.bind_all("<Control-MouseWheel>", self.on_mouse_wheel_zoom)

        run_updater(REPO_OWNER, REPO_NAME, CURRENT_VERSION)

        self.set_font_size(self.base_font_size)
        self.after(100, self.process_thread_queue)

    def change_appearance_mode(self, new_mode):
        ctk.set_appearance_mode(new_mode.lower())
        self.config_manager.config['APPEARANCE_MODE'] = new_mode
        self.config_manager.save_config()

    def process_thread_queue(self):
        try:
            task_name, data = self.thread_queue.get_nowait()

            if "update_ui" in task_name:
                if task_name == "update_ui_step1":
                    self.update_ui_after_step1(data)
                elif task_name == "update_ui_step2":
                    self.update_ui_after_step2(data)
                elif task_name == "update_ui_step3":
                    self.update_ui_after_step3(data)
            elif "recalculation_done" in task_name:
                self.update_ui_after_recalculation(data)
            elif "stabilization_proposal" in task_name:
                self.handle_stabilization_proposal(data)

            elif task_name == "export_done":
                messagebox.showinfo("내보내기 성공", f"계획이 {data}로 저장되었습니다.")
                self.update_status_bar("엑셀 내보내기 완료")
            elif task_name == "error":
                messagebox.showerror("작업 오류", data)
                self.update_status_bar("오류 발생")

            if not self.stabilization_active:
                self.set_ui_busy_state(False)

        except Empty:
            pass
        finally:
            self.after(100, self.process_thread_queue)

    def run_in_thread(self, worker_func, start_message):
        if self.is_task_running:
            messagebox.showwarning("작업 중", "이미 다른 작업이 실행 중입니다.")
            return

        self.set_ui_busy_state(True, start_message)
        thread = threading.Thread(target=worker_func, daemon=True)
        thread.start()

    def _animate_status_bar(self):
        animation_char = self.animation_chars[self.animation_idx % len(self.animation_chars)]
        self.status_bar.configure(text=f"현재 파일: {self.current_file} | 상태: {self.status_message_base}{animation_char}")
        self.animation_idx += 1
        self.animation_job = self.after(300, self._animate_status_bar)

    def set_ui_busy_state(self, is_busy, message=""):
        self.is_task_running = is_busy
        state = "disabled" if is_busy else "normal"
        cursor = "watch" if is_busy else ""

        self.step1_button.configure(state=state)
        self.step2_button.configure(state=state if self.current_step >=1 else "disabled")
        self.step3_button.configure(state=state if self.current_step >=2 else "disabled")
        self.step4_button.configure(state=state if self.current_step >=1 else "disabled")
        self.stabilize_button.configure(state=state if self.current_step >=2 else "disabled")
        self.save_settings_button.configure(state=state)
        self.daily_truck_button.configure(state=state)
        self.non_shipping_button.configure(state=state)
        self.safety_stock_button.configure(state=state)
        if hasattr(self, 'item_order_button'):
            self.item_order_button.configure(state=state)
        if hasattr(self, 'toggle_mode_button'):
            self.toggle_mode_button.configure(state=state if self.current_step >= 2 else "disabled")

        self.config(cursor=cursor)
        self.update_idletasks()

        if is_busy:
            self.status_message_base = message
            if self.animation_job:
                self.after_cancel(self.animation_job)
            self._animate_status_bar()
        else:
            if self.animation_job:
                self.after_cancel(self.animation_job)
                self.animation_job = None
            if not message:
                self.update_status_bar()

    def on_closing(self):
        try:
            self.config_manager.save_config()
            self.unbind_all("<Control-MouseWheel>")
            plt.close('all')
            if messagebox.askokcancel("종료", "프로그램을 종료하시겠습니까?"):
                self.destroy()
        except Exception as e:
            logging.error(f"Closing error: {e}")
            self.destroy()

    def set_font_size(self, size):
        size = max(8, min(40, size))
        self.base_font_size = size
        self.config_manager.config['FONT_SIZE'] = size

        self.font_normal.configure(size=size)
        self.font_header.configure(size=size + 1)
        self.font_big_bold.configure(size=size + 9)
        self.font_small.configure(size=size - 1)
        self.font_bold.configure(size=size)
        self.font_italic.configure(size=size)
        self.font_kpi.configure(size=size + 3)
        self.font_edit.configure(size=size)

        style = ttk.Style()
        style.configure("Treeview", rowheight=int(size * 2.2), font=self.font_normal)
        style.configure("Treeview.Heading", font=self.font_header)

        if hasattr(self, 'sidebar_title'):
            self.sidebar_title.configure(font=self.font_big_bold)
            self.step1_button.configure(font=self.font_normal)
            self.step2_button.configure(font=self.font_normal)
            self.step3_button.configure(font=self.font_normal)
            self.step4_button.configure(font=self.font_normal)
            self.stabilize_button.configure(font=self.font_normal)
            self.font_size_title_label.configure(font=self.font_normal)
            self.font_minus_button.configure(font=self.font_normal)
            self.font_size_label.configure(text=str(size), font=self.font_normal)
            self.font_plus_button.configure(font=self.font_normal)
            self.settings_frame.configure(label_font=self.font_bold)
            for label in self.setting_labels:
                label.configure(font=self.font_normal)
            for entry in self.settings_entries.values():
                entry.configure(font=self.font_normal)
            for cb in self.day_checkboxes.values():
                cb.configure(font=self.font_normal)
            self.daily_truck_button.configure(font=self.font_normal)
            self.non_shipping_button.configure(font=self.font_normal)
            self.safety_stock_button.configure(font=self.font_normal)
            self.item_order_button.configure(font=self.font_normal)
            self.save_settings_button.configure(font=self.font_normal)
            self.search_label.configure(font=self.font_normal)
            self.search_entry.configure(font=self.font_normal)
            self.toggle_mode_button.configure(font=self.font_normal)
            self.lbl_models_found.configure(font=self.font_kpi)
            self.lbl_total_quantity.configure(font=self.font_kpi)
            self.lbl_date_range.configure(font=self.font_kpi)
            self.detail_tab_title.configure(font=self.font_bold)
            self.status_bar.configure(font=self.font_normal)

        logging.info(f"폰트 크기를 {size}로 변경했습니다.")

    def change_font_size(self, delta):
        self.set_font_size(self.base_font_size + delta)

    def prompt_for_font_size(self, event=None):
        dialog = ctk.CTkInputDialog(text="새 폰트 크기를 입력하세요:", title="폰트 크기 변경")
        new_size_str = dialog.get_input()
        if new_size_str:
            try:
                new_size = int(new_size_str)
                self.set_font_size(new_size)
            except (ValueError, TypeError):
                messagebox.showerror("입력 오류", "유효한 숫자를 입력해주세요.", parent=self)

    def on_mouse_wheel_zoom(self, event):
        self.set_font_size(self.base_font_size + (1 if event.delta > 0 else -1))

    def toggle_sidebar(self):
        if self.sidebar_visible:
            self.paned_window.forget(self.sidebar_frame)
            self.sidebar_toggle_button.configure(text="▶")
            self.sidebar_visible = False
        else:
            self.paned_window.add(self.sidebar_frame, before=self.main_content_container)
            self.paned_window.sash_place(0, self.calculated_sidebar_width, 0)
            self.sidebar_toggle_button.configure(text="◀")
            self.sidebar_visible = True
    
    def create_widgets(self):
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)

        self.paned_window = PanedWindow(self, orient=HORIZONTAL, sashrelief=tk.RAISED, bg="#D3D3D3", opaqueresize=False)
        self.paned_window.grid(row=0, column=0, sticky="nsew")

        self.sidebar_frame = ctk.CTkFrame(self.paned_window, corner_radius=0)
        self.sidebar_frame.grid_rowconfigure(6, weight=1)

        self.main_content_container = ctk.CTkFrame(self.paned_window, fg_color="transparent")
        self.main_content_container.grid_rowconfigure(0, weight=1)
        self.main_content_container.grid_columnconfigure(1, weight=1)

        self.paned_window.add(self.sidebar_frame)
        self.paned_window.add(self.main_content_container)

        self.sidebar_toggle_button = ctk.CTkButton(self.main_content_container, text="◀", command=self.toggle_sidebar, width=20, height=40, corner_radius=5)
        self.sidebar_toggle_button.grid(row=0, column=0, sticky="w", pady=10)

        main_area_frame = ctk.CTkFrame(self.main_content_container, fg_color="transparent")
        main_area_frame.grid(row=0, column=1, sticky="nsew", padx=(10, 0))
        main_area_frame.grid_columnconfigure(0, weight=1)
        main_area_frame.grid_rowconfigure(1, weight=1) 

        self.sidebar_title = ctk.CTkLabel(self.sidebar_frame, text="PlanForge Pro", font=self.font_big_bold)
        self.sidebar_title.pack(pady=20)

        self.step1_button = ctk.CTkButton(self.sidebar_frame, text="1. 생산계획 불러오기", command=self.run_step1_aggregate, font=self.font_normal)
        self.step1_button.pack(fill='x', padx=20, pady=5)
        self.step2_button = ctk.CTkButton(self.sidebar_frame, text="2. 재고 반영", command=self.run_step2_simulation, state="disabled", font=self.font_normal)
        self.step2_button.pack(fill='x', padx=20, pady=5)
        self.step3_button = ctk.CTkButton(self.sidebar_frame, text="3. 수동 조정 적용", command=self.run_step3_adjustments, state="disabled", font=self.font_normal)
        self.step3_button.pack(fill='x', padx=20, pady=5)
        self.step4_button = ctk.CTkButton(self.sidebar_frame, text="4. 계획 내보내기 (Excel)", command=self.export_to_excel, state="disabled", font=self.font_normal)
        self.step4_button.pack(fill='x', padx=20, pady=5)
        self.stabilize_button = ctk.CTkButton(self.sidebar_frame, text="✨ 재고 안정화 실행", command=self.run_stabilization, state="disabled", font=self.font_normal, fg_color="#0B5345", hover_color="#117A65")
        self.stabilize_button.pack(fill='x', padx=20, pady=(15, 5))

        font_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        font_frame.pack(fill='x', padx=20, pady=10)
        self.font_size_title_label = ctk.CTkLabel(font_frame, text="폰트 크기:", font=self.font_normal)
        self.font_size_title_label.pack(side="left")
        self.font_minus_button = ctk.CTkButton(font_frame, text="-", width=30, command=lambda: self.change_font_size(-1), font=self.font_normal)
        self.font_minus_button.pack(side="left", padx=5)
        self.font_size_label = ctk.CTkLabel(font_frame, text=str(self.base_font_size), font=self.font_normal, cursor="hand2")
        self.font_size_label.pack(side="left")
        self.font_size_label.bind("<Button-1>", self.prompt_for_font_size)
        self.font_plus_button = ctk.CTkButton(font_frame, text="+", width=30, command=lambda: self.change_font_size(1), font=self.font_normal)
        self.font_plus_button.pack(side="left", padx=5)

        theme_frame = ctk.CTkFrame(self.sidebar_frame, fg_color="transparent")
        theme_frame.pack(fill='x', padx=20, pady=(5,10))
        ctk.CTkLabel(theme_frame, text="테마:", font=self.font_normal).pack(side="left")

        self.theme_button = ctk.CTkSegmentedButton(theme_frame,
                                                  values=["Light", "Dark", "System"],
                                                  command=self.change_appearance_mode,
                                                  font=self.font_normal)
        self.theme_button.pack(side="left", padx=10, expand=True, fill="x")
        self.theme_button.set(self.config_manager.config.get('APPEARANCE_MODE', 'System'))

        self.settings_frame = ctk.CTkScrollableFrame(self.sidebar_frame, label_text="시스템 설정", label_font=self.font_bold)
        self.settings_frame.pack(fill='both', expand=True, padx=15, pady=10)

        self.settings_entries = {}
        settings_map = {'팔레트당 수량': 'PALLET_SIZE', '리드타임 (일)': 'LEAD_TIME_DAYS', '트럭당 팔레트 수': 'PALLETS_PER_TRUCK', '기본 최대 차수': 'MAX_TRUCKS_PER_DAY'}
        self.setting_labels = []
        for label_text, key in settings_map.items():
            frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
            frame.pack(fill='x', pady=2, padx=5)
            label = ctk.CTkLabel(frame, text=label_text, width=120, anchor='w', font=self.font_normal)
            label.pack(side='left')
            self.setting_labels.append(label)
            entry = ctk.CTkEntry(frame, font=self.font_normal)
            entry.pack(side='left', fill='x', expand=True)
            self.settings_entries[key] = entry

        self.delivery_days_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        self.delivery_days_frame.pack(fill='x', pady=5, padx=5)
        ctk.CTkLabel(self.delivery_days_frame, text="납품 요일:", font=self.font_normal).pack(anchor="w", pady=(0, 2))
        checkbox_container = ctk.CTkFrame(self.delivery_days_frame, fg_color="transparent")
        checkbox_container.pack(fill='x')
        self.day_checkboxes = {}
        day_names = ["월", "화", "수", "목", "금", "토", "일"]
        for i, day in enumerate(day_names):
            state = self.config_manager.config.get('DELIVERY_DAYS', {}).get(str(i), 'False') == 'True'
            cb = ctk.CTkCheckBox(checkbox_container, text=day, onvalue=True, offvalue=False, font=self.font_normal, width=1)
            cb.grid(row=i // 4, column=i % 4, padx=2, pady=1, sticky='w')
            if state: cb.select()
            self.day_checkboxes[i] = cb

        self.daily_truck_button = ctk.CTkButton(self.settings_frame, text="일자별 차수/PL 설정", command=self.open_daily_truck_dialog, font=self.font_normal)
        self.daily_truck_button.pack(fill='x', padx=5, pady=5)
        self.non_shipping_button = ctk.CTkButton(self.settings_frame, text="휴무일/공휴일 설정", command=self.open_holiday_dialog, font=self.font_normal)
        self.non_shipping_button.pack(fill='x', padx=5, pady=5)
        self.safety_stock_button = ctk.CTkButton(self.settings_frame, text="품목별 최소 재고 설정", command=self.open_safety_stock_dialog, font=self.font_normal)
        self.safety_stock_button.pack(fill='x', padx=5, pady=5)
        self.item_order_button = ctk.CTkButton(self.settings_frame, text="품목 순서/규칙 변경", command=self.open_item_order_dialog, font=self.font_normal)
        self.item_order_button.pack(fill='x', padx=5, pady=5)

        path_frame = ctk.CTkFrame(self.settings_frame, fg_color="transparent")
        path_frame.pack(fill='x', padx=5, pady=(10, 5))
        ctk.CTkLabel(path_frame, text="자동 저장 경로:", font=self.font_normal).pack(anchor="w")
        self.auto_save_path_entry = ctk.CTkEntry(path_frame, textvariable=self.auto_save_path_var, state="disabled", font=self.font_small)
        self.auto_save_path_entry.pack(fill='x', pady=(0, 5))
        ctk.CTkButton(path_frame, text="경로 변경", command=self.change_auto_save_path, font=self.font_normal).pack(fill='x')

        self.save_settings_button = ctk.CTkButton(self.sidebar_frame, text="설정 저장 및 재계산", command=self.save_settings_and_recalculate, fg_color="#1F6AA5", font=self.font_normal)
        self.save_settings_button.pack(fill='x', padx=20, pady=10, side='bottom')
        self.load_settings_to_gui()

        top_container = ctk.CTkFrame(main_area_frame, fg_color="transparent")
        top_container.grid(row=0, column=0, sticky="new", pady=(0, 5))
        top_container.grid_columnconfigure(0, weight=1)

        search_and_toggle_frame = ctk.CTkFrame(top_container, fg_color="transparent")
        search_and_toggle_frame.grid(row=0, column=0, sticky="ew", pady=(0,5))
        search_and_toggle_frame.grid_columnconfigure(1, weight=1)

        self.search_label = ctk.CTkLabel(search_and_toggle_frame, text="품목 검색:", font=self.font_normal)
        self.search_label.grid(row=0, column=0, sticky="w", padx=(0,5))
        self.search_entry = ctk.CTkEntry(search_and_toggle_frame, font=self.font_normal)
        self.search_entry.grid(row=0, column=1, sticky="ew")
        self.search_entry.bind("<KeyRelease>", self.filter_grid)
        
        self.toggle_mode_button = ctk.CTkButton(search_and_toggle_frame, text="조정치 보기", command=self.toggle_display_mode, font=self.font_normal, width=120, state="disabled")
        self.toggle_mode_button.grid(row=0, column=2, sticky="e", padx=(10,0))
        
        self.kpi_frame = ctk.CTkFrame(top_container, fg_color="#EAECEE", corner_radius=5)
        self.kpi_frame.grid(row=1, column=0, sticky="ew", pady=(5,0))
        self.kpi_frame.grid_columnconfigure((0,1,2), weight=1)
        self.lbl_models_found = ctk.CTkLabel(self.kpi_frame, text="처리된 모델 수: -", font=self.font_kpi)
        self.lbl_models_found.grid(row=0, column=0, padx=10, pady=10)
        self.lbl_total_quantity = ctk.CTkLabel(self.kpi_frame, text="총생산량: -", font=self.font_kpi)
        self.lbl_total_quantity.grid(row=0, column=1, padx=10, pady=10)
        self.lbl_date_range = ctk.CTkLabel(self.kpi_frame, text="계획 기간: -", font=self.font_kpi)
        self.lbl_date_range.grid(row=0, column=2, padx=10, pady=10)
        self.kpi_frame.grid_remove() 

        self.warnings_header_frame = ctk.CTkFrame(top_container, fg_color="transparent")
        self.warnings_header_frame.grid(row=2, column=0, sticky="ew", pady=(5,0))
        self.warnings_toggle_button = ctk.CTkButton(self.warnings_header_frame, text="🚨 시스템 경고", font=self.font_bold, anchor="w", command=self.toggle_warnings_details)
        self.warnings_toggle_button.pack(fill="x", expand=True)
        self.warnings_header_frame.grid_remove() 

        self.warnings_container = ctk.CTkFrame(top_container, fg_color="transparent")
        self.warnings_container.grid(row=3, column=0, sticky="ew", pady=5)
        self.warnings_container.grid_columnconfigure(0, weight=1)
        self.warnings_container.grid_columnconfigure(1, weight=1)

        self.unmet_demand_frame = ctk.CTkFrame(self.warnings_container, fg_color="#FFDDE1", corner_radius=5)
        self.unmet_demand_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 5))
        self.unmet_list_frame = ctk.CTkScrollableFrame(self.unmet_demand_frame, label_text="계획 실패", label_font=self.font_bold, label_text_color="#C0392B")
        self.unmet_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.shortage_frame = ctk.CTkFrame(self.warnings_container, fg_color="#FFF5E1", corner_radius=5)
        self.shortage_frame.grid(row=0, column=1, sticky="nsew", padx=(5, 0))
        self.shortage_list_frame = ctk.CTkScrollableFrame(self.shortage_frame, label_text="재고 부족", label_font=self.font_bold, label_text_color="#E67E22")
        self.shortage_list_frame.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.warnings_container.grid_remove() 

        self.tabview = ctk.CTkTabview(main_area_frame)
        self.tabview.grid(row=1, column=0, sticky="nsew") 
        self.master_tab = self.tabview.add("개요")
        self.detail_tab = self.tabview.add("상세")
        self.master_tab.grid_columnconfigure(0, weight=1)
        self.master_tab.grid_rowconfigure(0, weight=1)
        self.detail_tab.grid_columnconfigure(0, weight=1)
        self.detail_tab.grid_rowconfigure(1, weight=1)

        style = ttk.Style()
        style.configure("Treeview", rowheight=int(self.base_font_size * 2.2), font=self.font_normal)
        style.configure("Treeview.Heading", font=self.font_header)

        tree_container = ctk.CTkFrame(self.master_tab, fg_color="transparent")
        tree_container.grid(row=0, column=0, sticky="nsew")
        tree_container.grid_rowconfigure(0, weight=1)
        tree_container.grid_columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tree_container, style="Treeview", show="headings")
        self.tree.grid(row=0, column=0, sticky="nsew")

        v_scrollbar = ctk.CTkScrollbar(tree_container, orientation="vertical", command=self.tree.yview)
        v_scrollbar.grid(row=0, column=1, sticky="ns")
        h_scrollbar = ctk.CTkScrollbar(tree_container, orientation="horizontal", command=self.tree.xview)
        h_scrollbar.grid(row=1, column=0, sticky="ew")
        self.tree.configure(yscrollcommand=v_scrollbar.set, xscrollcommand=h_scrollbar.set)

        self.tree.tag_configure('highlight', background='#D6EAF8', foreground='#154360')
        self.tree.tag_configure('locked', background='#D5F5E3', foreground="#0E6655")
        self.tree.tag_configure('total', font=self.font_bold, background='#EAECEE')
        self.tree.tag_configure('separator', background='#FFFFFF')
        self.tree.tag_configure('optimized', foreground='#1A5276')

        self.tree.bind("<Double-1>", self.on_treeview_double_click)
        self.tree.bind("<Button-3>", self.on_treeview_right_click)

        self.detail_tab_title = ctk.CTkLabel(self.detail_tab, text="상세: 선택된 모델의 출고 시뮬레이션", font=self.font_bold)
        self.detail_tab_title.grid(row=0, column=0, sticky="w", padx=10, pady=(5,0))
        self.detail_frame = ctk.CTkScrollableFrame(self.detail_tab, label_text="")
        self.detail_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=5)
        self.status_bar = ctk.CTkLabel(self, text="준비 완료", anchor="w", font=self.font_normal)
        self.status_bar.grid(row=1, column=0, sticky="ew", padx=10, pady=(0, 5))

        self.update_idletasks()
        padding = 40
        self.calculated_sidebar_width = self.sidebar_frame.winfo_reqwidth() + padding
        logging.info(f"계산된 사이드바 너비: {self.calculated_sidebar_width}")

        self.paned_window.paneconfigure(self.sidebar_frame, minsize=self.calculated_sidebar_width)
        self.paned_window.sash_place(0, self.calculated_sidebar_width, 0)
    
    def toggle_display_mode(self):
        if self.display_mode == 'sum':
            self.display_mode = 'adjustment'
            self.toggle_mode_button.configure(text="합계 보기")
        else:
            self.display_mode = 'sum'
            self.toggle_mode_button.configure(text="조정치 보기")
        
        self.filter_grid() 

    def toggle_warnings_details(self):
        self.warnings_visible = not self.warnings_visible
        self.update_all_warnings_ui()

    def on_treeview_double_click(self, event):
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return

        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or not column_id: return

        tags = self.tree.item(item_id, "tags")
        if 'total' in tags or 'separator' in tags: return

        model_name = self.tree.item(item_id, "values")[0]

        if self.tree.column(column_id, "id") == "Item Code":
            self.on_row_double_click(model_name)
        elif self.current_step >= 2:
            self.edit_shipment_value(item_id, column_id)

    def on_treeview_right_click(self, event):
        if self.current_step < 2: return

        item_id = self.tree.identify_row(event.y)
        column_id = self.tree.identify_column(event.x)
        if not item_id or not column_id or self.tree.column(column_id, "id") == "Item Code":
            return

        tags = self.tree.item(item_id, "tags")
        if 'total' in tags or 'separator' in tags: return

        model, date, truck_num, _ = self.get_cell_info(item_id, column_id)
        if not all([model, date, truck_num is not None]): return

        is_locked = any(s['model'] == model and s['date'] == date and s['truck_num'] == truck_num for s in self.processor.fixed_shipments)

        context_menu = Menu(self, tearoff=0)
        if is_locked:
            context_menu.add_command(label="출고량 고정 해제", command=lambda: self.unfix_shipment(model, date, truck_num))
        else:
            context_menu.add_command(label="출고량 고정 (Lock)", command=lambda: self.fix_shipment(model, date, truck_num, from_menu=True))

        context_menu.add_separator()
        context_menu.add_command(label="수량 직접 수정", command=lambda: self.edit_shipment_value(item_id, column_id))

        context_menu.tk_popup(event.x_root, event.y_root)

    def populate_treeview(self, df_to_show):
        for item in self.tree.get_children():
            self.tree.delete(item)

        if df_to_show.empty:
            self.tree["columns"] = []
            return

        def get_item_code_width(df):
            header_font = tkfont.Font(font=self.font_header)
            content_font = tkfont.Font(font=self.font_normal)
            header_width = header_font.measure("품목 코드")
            max_content_width = 0
            if not df.empty and df.index.size > 0:
                max_content_width = max(content_font.measure(str(s)) for s in df.index)
            return max(header_width, max_content_width) + 30

        if self.current_step < 2:
            start_date_to_show = self.processor.planning_start_date or self.processor.date_cols[0].date()
            display_cols = [d for d in self.processor.date_cols if d in df_to_show.columns and df_to_show[d].sum() > 0 and d.date() >= start_date_to_show]
            headers = ["Item Code"] + [d.strftime('%m-%d') for d in display_cols]
            self.tree["columns"] = headers
            
            item_code_width = get_item_code_width(df_to_show)
            self.tree.column("Item Code", anchor="w", width=item_code_width, minwidth=item_code_width, stretch=False)
            self.tree.heading("Item Code", text="품목 코드", anchor="w")

            for col in headers[1:]:
                self.tree.column(col, anchor="e", width=90, minwidth=90, stretch=False)
                self.tree.heading(col, text=col, anchor="e")

            for index, row in df_to_show.iterrows():
                values = [index] + [f"{int(row.get(d, 0)):,}" for d in display_cols]
                tags = ('highlight',) if index in self.processor.highlight_models else ()
                self.tree.insert("", "end", values=values, tags=tags)

            if not df_to_show.empty:
                pallet_size = self.processor.config.get('PALLET_SIZE', 60)
                total_pcs_values = ["합계 (PCS)"]
                total_pallets_values = ["합계 (Pallet)"]
                for col_date in display_cols:
                    total_pcs = df_to_show[col_date].sum()
                    total_pcs_values.append(f"{int(total_pcs):,}")
                    pallet_series = np.ceil(df_to_show[col_date] / pallet_size).where(df_to_show[col_date] > 0, 0)
                    total_pallets = pallet_series.sum()
                    total_pallets_values.append(f"{int(total_pallets):,}")
                self.tree.insert("", "end", values=[""]*len(headers), tags=('separator',))
                self.tree.insert("", "end", values=total_pcs_values, tags=('total',))
                self.tree.insert("", "end", values=total_pallets_values, tags=('total',))
        else:
            all_shipment_cols_raw = [c for c in df_to_show.columns if isinstance(c, str) and c.startswith('출고_') and df_to_show[c].sum() > 0]
            def sort_key(col_name):
                parts = col_name.split('_')
                date_str = parts[2]
                truck_num = int(re.search(r'\d+', parts[1]).group())
                return (date_str, truck_num)
            all_shipment_cols = sorted(all_shipment_cols_raw, key=sort_key)
            headers = ["Item Code"]
            header_map = {}
            for col_name in all_shipment_cols:
                _, truck, date = col_name.split('_')
                header_text = f"{date[:2]}-{date[2:]} {truck}"
                headers.append(header_text)
                header_map[header_text] = col_name
            self.tree["columns"] = headers

            item_code_width = get_item_code_width(df_to_show)
            self.tree.column("Item Code", anchor="w", width=item_code_width, minwidth=item_code_width, stretch=False)
            self.tree.heading("Item Code", text="품목 코드", anchor="w")

            for h in headers[1:]:
                self.tree.column(h, anchor="e", width=130, minwidth=130, stretch=False)
                self.tree.heading(h, text=h, anchor="e")

            for index, row in df_to_show.iterrows():
                tags_list = ['highlight'] if index in self.processor.highlight_models else []
                is_row_optimized = False
                values = [index]
                for h in headers[1:]:
                    total_qty = int(row.get(header_map[h], 0))
                    model, date, truck_num = self.get_cell_info_from_header(index, h)
                    
                    cell_value = f"{total_qty:,}"
                    if self.display_mode == 'adjustment':
                        opt_key = (date, truck_num, model)
                        if opt_key in self.processor.optimized_additions:
                            added_qty = self.processor.optimized_additions[opt_key]
                            original_qty = total_qty - added_qty
                            if added_qty > 0:
                                cell_value = f"{original_qty:,} (+{added_qty:,})"
                                is_row_optimized = True

                    values.append(cell_value)
                    
                    if any(s['model'] == model and s['date'] == date and s['truck_num'] == truck_num for s in self.processor.fixed_shipments):
                        if 'locked' not in tags_list: tags_list.append('locked')
                
                if is_row_optimized:
                    tags_list.append('optimized')

                self.tree.insert("", "end", values=values, tags=tuple(tags_list))

            if not df_to_show.empty:
                pallet_size = self.processor.config.get('PALLET_SIZE', 60)
                total_pcs_values = ["합계 (PCS)"]
                total_pallets_values = ["합계 (Pallet)"]
                for h in headers[1:]:
                    col_name = header_map[h]
                    total_pcs = df_to_show[col_name].sum()
                    total_pcs_values.append(f"{int(total_pcs):,}")
                    
                    total_pallets = 0
                    if pallet_size > 0:
                        for item_code in df_to_show.index:
                            item_qty = df_to_show.loc[item_code, col_name]
                            if item_qty > 0:
                                total_pallets += math.ceil(item_qty / pallet_size)
                                
                    total_pallets_values.append(f"{int(total_pallets):,}")
                self.tree.insert("", "end", values=[""]*len(headers), tags=('separator',))
                self.tree.insert("", "end", values=total_pcs_values, tags=('total',))
                self.tree.insert("", "end", values=total_pallets_values, tags=('total',))

    def filter_grid(self, event=None):
        if self.processor is None: return
        self.processor._ensure_item_master_loaded()

        df_source = self.processor.aggregated_plan_df if self.current_step < 2 else self.processor.simulated_plan_df
        if df_source is None:
            df_to_show = pd.DataFrame()
        else:
            sum_cols = []
            if self.current_step < 2 and self.processor.date_cols:
                sum_cols = self.processor.date_cols
            elif self.current_step >= 2:
                sum_cols = [c for c in df_source.columns if isinstance(c, str) and c.startswith('출고_')]

            if sum_cols:
                df_to_show = df_source.loc[df_source[sum_cols].sum(axis=1) > 0].copy()
            else:
                df_to_show = df_source.copy()

        search_term = self.search_entry.get().lower()
        if search_term:
            df_to_show = df_to_show[df_to_show.index.str.lower().str.contains(search_term)]

        self.populate_treeview(df_to_show)

    def run_step1_aggregate(self):
        file_path = filedialog.askopenfilename(title="생산계획 엑셀 파일 선택", filetypes=(("Excel", "*.xlsx *.xls"),))
        if not file_path: return

        def worker():
            try:
                self.processor.current_filepath = file_path
                self.processor.process_plan_file()
                self.current_file = os.path.basename(file_path)
                df = self.processor.aggregated_plan_df
                if df is None or df.empty:
                    self.thread_queue.put(("error", "처리할 생산 계획 데이터가 없습니다."))
                    return
                plan_cols = self.processor.date_cols
                df_filtered = df[df[plan_cols].sum(axis=1) > 0]
                result = { "models_found": len(df_filtered.index), "total_qty": df_filtered[plan_cols].sum().sum(), "date_range": f"{plan_cols[0].strftime('%y/%m/%d')} ~ {plan_cols[-1].strftime('%y/%m/%d')}" }
                self.thread_queue.put(("update_ui_step1", result))
            except Exception as e:
                self.thread_queue.put(("error", f"1단계 파일 처리 실패: {e}"))

        self.run_in_thread(worker, "생산계획 집계 중")

    def run_step2_simulation(self):
        if self.current_step < 1: return

        dialog_inv = InventoryInputDialog(self, font_normal=self.font_normal)
        self.wait_window(dialog_inv)
        inventory_input = dialog_inv.result
        if not inventory_input: return
        
        input_type, input_data = inventory_input
        try:
            if input_type == 'text':
                self.inventory_text_backup = input_data
                self.processor.load_inventory_from_text(input_data)
            else: 
                self.processor.load_inventory_from_file(input_data)
                self.inventory_text_backup = None
        except Exception as e:
            messagebox.showerror("오류", f"재고 데이터 처리 중 오류가 발생했습니다: {e}", parent=self)
            return

        available_dates = sorted([d.date() for d in self.processor.date_cols])
        dialog_last_ship = LastShipmentDialog(self, available_dates, font_normal=self.font_normal)
        self.wait_window(dialog_last_ship)
        last_shipment_date = dialog_last_ship.result
        if last_shipment_date is None: return 
        
        if last_shipment_date == "None":
            self.processor.planning_start_date = self.processor.inventory_date
        else:
            self.processor.planning_start_date = last_shipment_date + timedelta(days=1)

        dialog_in_transit = InTransitDialog(self, self.processor.allowed_models, self.processor.config['LEAD_TIME_DAYS'], self.processor.inventory_date, font_normal=self.font_normal)
        self.wait_window(dialog_in_transit)
        self.processor.in_transit_inventory = dialog_in_transit.result

        def worker():
            try:
                self.processor.run_simulation(in_transit_inventory=self.processor.in_transit_inventory)
                
                df = self.processor.simulated_plan_df
                if df is None or df.empty:
                    self.thread_queue.put(("error", "시뮬레이션 후 데이터가 없습니다."))
                    return
                
                ship_cols = [c for c in df.columns if isinstance(c, str) and c.startswith('출고_')]
                total_ship = df[ship_cols].sum().sum()
                result = {"total_ship": total_ship}
                self.thread_queue.put(("update_ui_step2", result))
            except Exception as e:
                logging.exception("2단계 시뮬레이션 중 오류 발생")
                self.thread_queue.put(("error", f"2단계 시뮬레이션 실패: {e}"))

        self.run_in_thread(worker, "재고 반영 및 시뮬레이션 중")

    def save_step1_result_to_excel(self):
        try:
            df = self.processor.aggregated_plan_df
            plan_cols = self.processor.date_cols

            if df is None or df.empty or not plan_cols:
                logging.info("Excel 저장을 건너뜁니다: 저장할 데이터가 없습니다.")
                return

            df_filtered = df[df[plan_cols].sum(axis=1) > 0]
            if df_filtered.empty:
                logging.info("Excel 저장을 건너뜁니다: 필터링 후 데이터가 없습니다.")
                return

            base_path = self.config_manager.config.get('AUTO_SAVE_PATH', os.path.join(os.path.expanduser('~'), 'Downloads'))
            output_dir = os.path.join(base_path, '생산계획_집계결과')
            os.makedirs(output_dir, exist_ok=True)
            timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            base_excel_name = os.path.splitext(self.current_file)[0]
            filename = f"집계_{base_excel_name}_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)

            plan_df = df_filtered[[c for c in plan_cols if c in df_filtered.columns]].copy()
            plan_df.columns = pd.MultiIndex.from_tuples(
                [(d.strftime('%m-%d'), '생산량') for d in plan_df.columns],
                names=['날짜', '구분']
            )

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                plan_df.to_excel(writer, sheet_name='생산 계획 집계')

                workbook = writer.book
                worksheet = writer.sheets['생산 계획 집계']

                max_length = max(len(str(s)) for s in plan_df.index) if not plan_df.empty else 10
                worksheet.column_dimensions['A'].width = max_length + 5

                blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                header_rows = 2
                for r_idx, model_name in enumerate(plan_df.index):
                    if model_name in self.processor.highlight_models:
                        row_to_format = worksheet[r_idx + 1 + header_rows]
                        for cell in row_to_format:
                            cell.fill = blue_fill

                if not plan_df.empty:
                    pallet_size = self.processor.config.get('PALLET_SIZE', 60)
                    total_pcs = plan_df.sum()
                    pallet_df = np.ceil(plan_df / pallet_size).where(plan_df > 0, 0)
                    total_pallets = pallet_df.sum().astype(int)
                    bold_font = Font(bold=True)
                    summary_start_row = len(plan_df) + 4

                    pcs_label_cell = worksheet.cell(row=summary_start_row, column=1, value="합계 (PCS)")
                    pcs_label_cell.font = bold_font
                    for i, total in enumerate(total_pcs):
                        cell = worksheet.cell(row=summary_start_row, column=i + 2, value=total)
                        cell.font = bold_font

                    pallet_label_cell = worksheet.cell(row=summary_start_row + 1, column=1, value="합계 (Pallet)")
                    pallet_label_cell.font = bold_font
                    for i, total in enumerate(total_pallets):
                        cell = worksheet.cell(row=summary_start_row + 1, column=i + 2, value=total)
                        cell.font = bold_font

                thin_border_side = Side(border_style="thin", color="000000")
                thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                    for cell in row:
                        cell.border = thin_border

            status_message = f"1단계 완료. 집계 결과가 엑셀로 저장되었습니다."
            self.update_status_bar(status_message)
            logging.info(f"생산 계획 집계 결과가 Excel로 저장되었습니다: {filepath}")

        except Exception as e:
            logging.error(f"Excel 저장 중 오류 발생: {e}", exc_info=True)
            messagebox.showwarning("저장 오류", f"집계 결과를 Excel로 저장하는 데 실패했습니다:\n{e}", parent=self)
            self.update_status_bar("1단계 완료 (Excel 저장 실패)")

    def update_ui_after_step1(self, data):
        self.current_step = 1
        self.processor.planning_start_date = None
        self.step2_button.configure(state="normal")
        self.step3_button.configure(state="disabled")
        self.step4_button.configure(state="disabled")
        self.stabilize_button.configure(state="disabled")
        self.toggle_mode_button.configure(state="disabled")
        [widget.destroy() for widget in self.detail_frame.winfo_children()]
        self.filter_grid()

        self.save_step1_result_to_excel()
        
        self.kpi_frame.grid_remove()
        self.warnings_header_frame.grid_remove()
        self.warnings_container.grid_remove()
        logging.info("1단계 완료. UI 업데이트 완료.")

    def update_ui_after_step2(self, data):
        self.kpi_frame.grid()
        self.current_step = 2
        self.step3_button.configure(state="normal")
        self.step4_button.configure(state="normal")
        self.stabilize_button.configure(state="normal")
        self.toggle_mode_button.configure(state="normal")
        self.lbl_models_found.configure(text=f"처리된 모델 수: {len(self.processor.simulated_plan_df)} 개")
        self.lbl_total_quantity.configure(text=f"총출고량: {data['total_ship']:,.0f} 개")

        sim_start_date = self.processor.planning_start_date
        plan_end_date = self.processor.date_cols[-1].date()
        self.lbl_date_range.configure(text=f"계획 기간: {sim_start_date.strftime('%y/%m/%d')} ~ {plan_end_date.strftime('%y/%m/%d')}")

        [widget.destroy() for widget in self.detail_frame.winfo_children()]
        self.filter_grid()
        self.update_status_bar("2단계: 출고 계획 시뮬레이션 완료.")
        self.check_shipment_capacity()
        
        self.update_all_warnings_ui()
        logging.info("2단계 완료. 시뮬레이션 결과 UI 업데이트 완료.")

    def run_step3_adjustments(self):
        self.processor._ensure_item_master_loaded()
        dialog = AdjustmentDialog(self, models=self.processor.allowed_models, font_normal=self.font_normal)
        self.wait_window(dialog)
        all_adjustments = dialog.result
        if all_adjustments is None: return

        def worker():
            try:
                self.processor.adjustments = [adj for adj in all_adjustments if adj['type'] in ['재고', '수요']]
                self.processor.fixed_shipment_reqs = [adj for adj in all_adjustments if adj['type'] == '고정 출고']
                if self.inventory_text_backup:
                    self.processor.load_inventory_from_text(self.inventory_text_backup)
                self.processor.run_simulation(adjustments=self.processor.adjustments, fixed_shipments=self.processor.fixed_shipments, fixed_shipment_reqs=self.processor.fixed_shipment_reqs, in_transit_inventory=self.processor.in_transit_inventory)
                total_ship = self.processor.simulated_plan_df[[c for c in self.processor.simulated_plan_df.columns if isinstance(c, str) and c.startswith('출고_')]].sum().sum()
                self.thread_queue.put(("update_ui_step3", {"total_ship": total_ship}))
            except Exception as e:
                self.thread_queue.put(("error", f"3단계 조정 실패: {e}"))

        self.run_in_thread(worker, "수동 조정 적용 및 재계산 중")

    def update_ui_after_step3(self, data):
        self.current_step = 3
        self.lbl_total_quantity.configure(text=f"총출고량: {data['total_ship']:,.0f} 개")
        [widget.destroy() for widget in self.detail_frame.winfo_children()]
        self.filter_grid()
        self.update_status_bar("3단계: 수동 조정 적용 완료.")
        self.check_shipment_capacity()
        
        self.update_all_warnings_ui()
        logging.info("3단계 완료. 조정 결과 UI 업데이트 완료.")

    def run_stabilization(self):
        if self.current_step < 2: return

        if not messagebox.askyesno("재고 안정화 실행",
                                "자동으로 계획 실패 및 재고 부족을 해결하는 최적의 계획을 찾습니다.\n\n"
                                "이 작업은 여러 번의 재계산을 포함하며, '일자별 차수/PL 설정'이 변경될 수 있습니다.\n"
                                "계속하시겠습니까?", parent=self):
            return

        self.applied_fixes = set() # 안정화 시작 시, 시도한 해결책 목록 초기화
        self.stabilization_active = True
        self.stabilization_iteration = 0
        self.run_stabilization_step()

    def run_stabilization_step(self):
        # self.stabilization_iteration += 1 # 🔴 이 줄은 handle_stabilization_proposal로 이동되었으므로 주석 처리합니다.
        if self.stabilization_iteration > 30:
            messagebox.showerror("안정화 실패", "최대 반복 횟수(30회)를 초과했습니다. 계획을 검토해주세요.")
            self.stabilization_active = False
            self.set_ui_busy_state(False)
            return

        def worker():
            try:
                if self.inventory_text_backup: self.processor.load_inventory_from_text(self.inventory_text_backup)
                self.processor.run_simulation(
                    adjustments=self.processor.adjustments,
                    fixed_shipments=self.processor.fixed_shipments,
                    fixed_shipment_reqs=self.processor.fixed_shipment_reqs,
                    in_transit_inventory=self.processor.in_transit_inventory
                )

                proposal = self.processor.find_stabilization_proposal()
                self.thread_queue.put(("stabilization_proposal", proposal))
            except Exception as e:
                self.thread_queue.put(("error", f"재고 안정화 중 오류 발생: {e}"))
                self.stabilization_active = False

        self.run_in_thread(worker, f"재고 안정화 실행 중 ({self.stabilization_iteration + 1}차 재계산)")

    def handle_stabilization_proposal(self, proposal):
        if proposal['type'] == 'error':
            messagebox.showerror("안정화 실패", proposal['message'])
            self.stabilization_active = False
            self.set_ui_busy_state(False)

        elif proposal['type'] == 'stable':
            messagebox.showinfo("안정화 완료", f"{proposal['message']} ({self.stabilization_iteration}회 반복)")
            self.stabilization_active = False
            self.set_ui_busy_state(False)
            self.update_ui_after_recalculation("재고 안정화 완료.")

        elif proposal['type'] == 'proposal':
            date = proposal['date']
            new_count = proposal['new_truck_count']
            reason = proposal['reason']
            
            # 제안된 해결책이 이전에 이미 시도했던 것인지 확인
            fix_signature = (date, new_count)
            if fix_signature in self.applied_fixes:
                messagebox.showerror("안정화 중단", 
                                   f"동일한 문제({reason})에 대해 이미 시도했던 해결책( {date.strftime('%m-%d')} {new_count}차로 증차)을 다시 제안했습니다.\n\n"
                                   "이는 근본적으로 재고가 부족하여 해결할 수 없는 문제입니다. '최소 재고 설정'을 조정하거나 수동으로 계획을 수정해주세요.")
                self.stabilization_active = False
                self.set_ui_busy_state(False)
                return
            
            self.applied_fixes.add(fix_signature)

            user_response = messagebox.askyesno(
                "안정화 제안",
                f"[{reason}] 문제를 해결하기 위해\n"
                f"{date.strftime('%Y-%m-%d')}의 최대 차수를 {new_count}회로 늘릴까요?",
                parent=self
            )
            
            # --- 🟢 수정된 부분 시작 ---
            if user_response:
                self.processor.config['DAILY_TRUCK_OVERRIDES'][date] = new_count
                logging.info(f"사용자 동의: {date} 차수를 {new_count}로 증가. (사유: {reason})")

                # UI는 계속 바쁜 상태로 두고, 다음 계산 스레드만 직접 시작
                self.stabilization_iteration += 1
                self.set_ui_busy_state(True, f"재고 안정화 실행 중 ({self.stabilization_iteration}차 재계산)")

                def worker():
                    try:
                        if self.inventory_text_backup: self.processor.load_inventory_from_text(self.inventory_text_backup)
                        self.processor.run_simulation(
                            adjustments=self.processor.adjustments,
                            fixed_shipments=self.processor.fixed_shipments,
                            fixed_shipment_reqs=self.processor.fixed_shipment_reqs,
                            in_transit_inventory=self.processor.in_transit_inventory
                        )
                        new_proposal = self.processor.find_stabilization_proposal()
                        self.thread_queue.put(("stabilization_proposal", new_proposal))
                    except Exception as e:
                        self.thread_queue.put(("error", f"재고 안정화 중 오류 발생: {e}"))
                        self.stabilization_active = False

                thread = threading.Thread(target=worker, daemon=True)
                thread.start()
            # --- 수정된 부분 끝 ---
            else:
                messagebox.showinfo("안정화 중단", "사용자가 제안을 거부하여 안정화 프로세스를 중단합니다.")
                self.stabilization_active = False
                self.set_ui_busy_state(False)

    def recalculate_with_fixed_values(self):
        def worker():
            try:
                if self.inventory_text_backup:
                    self.processor.load_inventory_from_text(self.inventory_text_backup)
                self.processor.run_simulation(adjustments=self.processor.adjustments, fixed_shipments=self.processor.fixed_shipments, fixed_shipment_reqs=self.processor.fixed_shipment_reqs, in_transit_inventory=self.processor.in_transit_inventory)
                self.thread_queue.put(("recalculation_done", "고정값 적용 및 재계산 완료"))
            except Exception as e:
                self.thread_queue.put(("error", f"재계산 실패: {e}"))
        self.run_in_thread(worker, "고정값 적용하여 재계산 중")

    def save_settings_and_recalculate(self):
        new_config = self.config_manager.config.copy()
        try:
            for key, entry_widget in self.settings_entries.items():
                new_config[key] = int(entry_widget.get())
            new_delivery_days = {str(i): str(self.day_checkboxes[i].get()) for i in range(7)}
            new_config['DELIVERY_DAYS'] = new_delivery_days
        except Exception as e:
            messagebox.showerror("설정 오류", f"설정값 저장 중 오류 발생: {e}"); return

        def worker():
            try:
                self.processor.config = new_config
                self.config_manager.config = new_config
                self.config_manager.save_config()
                if self.current_step >= 1:
                    if self.processor.aggregated_plan_df is not None:
                        if self.current_step >= 2:
                            if self.inventory_text_backup: self.processor.load_inventory_from_text(self.inventory_text_backup)
                            self.processor.run_simulation(adjustments=self.processor.adjustments, fixed_shipments=self.processor.fixed_shipments, fixed_shipment_reqs=self.processor.fixed_shipment_reqs, in_transit_inventory=self.processor.in_transit_inventory)
                self.thread_queue.put(("recalculation_done", "설정 저장 및 재계산 완료"))
            except Exception as e:
                self.thread_queue.put(("error", f"재계산 실패: {e}"))
        self.run_in_thread(worker, "설정 저장 및 재계산 중")

    def update_ui_after_recalculation(self, message):
        self.load_settings_to_gui()
        self.filter_grid()
        if self.last_selected_model:
            self.populate_detail_view(self.last_selected_model)
        
        self.update_all_warnings_ui()
        self.update_status_bar(message)
        messagebox.showinfo("성공", message)
        logging.info(message)

    def export_to_excel(self):
        if self.current_step < 2:
            messagebox.showwarning("오류", "먼저 '재고 반영 및 계획 시뮬레이션'을 실행해야 합니다.")
            return

        start_date = self.processor.planning_start_date.strftime('%m-%d')
        end_date = self.processor.date_cols[-1].strftime('%m-%d')
        filename = f"{start_date}~{end_date} 출고계획.xlsx"
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", initialfile=filename, filetypes=(("Excel", "*.xlsx"),))
        if not file_path: return

        def worker():
            try:
                df = self.processor.simulated_plan_df
                shipment_cols = [c for c in df.columns if isinstance(c, str) and c.startswith('출고_')]

                multi_index_cols = []
                for col_name in shipment_cols:
                    parts = col_name.split('_')
                    truck_num_str = parts[1]
                    date_str = parts[2]
                    formatted_date = f"{date_str[:2]}-{date_str[2:]}"
                    multi_index_cols.append((formatted_date, truck_num_str))

                shipment_df = df[shipment_cols].copy()
                shipment_df.columns = pd.MultiIndex.from_tuples(multi_index_cols, names=['날짜', '차수'])

                shipment_df = shipment_df.loc[:, shipment_df.sum() > 0]
                shipment_df = shipment_df.loc[shipment_df.sum(axis=1) > 0]

                sorted_models = self.processor.item_master_df.index
                shipment_df = shipment_df.reindex(index=sorted_models).dropna(how='all')

                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    shipment_df.to_excel(writer, sheet_name='출고 계획')

                    workbook = writer.book
                    worksheet = writer.sheets['출고 계획']

                    max_length = 0
                    for cell in worksheet['A']:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 5)
                    worksheet.column_dimensions['A'].width = adjusted_width

                    blue_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
                    header_rows = 2
                    for r_idx, model_name in enumerate(shipment_df.index):
                        if model_name in self.processor.highlight_models:
                            row_to_format = worksheet[r_idx + 1 + header_rows]
                            for cell in row_to_format:
                                cell.fill = blue_fill

                    if not shipment_df.empty:
                        pallet_size = self.processor.config.get('PALLET_SIZE', 60)

                        total_pcs = shipment_df.sum()

                        pallet_df = np.ceil(shipment_df / pallet_size).where(shipment_df > 0, 0)
                        total_pallets = pallet_df.sum().astype(int)

                        bold_font = Font(bold=True)

                        summary_start_row = len(shipment_df) + 4

                        pcs_label_cell = worksheet.cell(row=summary_start_row, column=1, value="합계 (PCS)")
                        pcs_label_cell.font = bold_font
                        for i, total in enumerate(total_pcs):
                            cell = worksheet.cell(row=summary_start_row, column=i + 2, value=total)
                            cell.font = bold_font

                        pallet_label_cell = worksheet.cell(row=summary_start_row + 1, column=1, value="합계 (Pallet)")
                        pallet_label_cell.font = bold_font
                        for i, total in enumerate(total_pallets):
                            cell = worksheet.cell(row=summary_start_row + 1, column=i + 2, value=total)
                            cell.font = bold_font

                    thin_border_side = Side(border_style="thin", color="000000")
                    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1, max_col=worksheet.max_column):
                        for cell in row:
                            cell.border = thin_border

                self.thread_queue.put(("export_done", file_path))
            except Exception as e:
                logging.exception("엑셀 내보내기 중 오류 발생")
                self.thread_queue.put(("error", f"내보내기 실패: {e}"))

        self.run_in_thread(worker, f"'{os.path.basename(file_path)}' 파일로 내보내는 중")

    def on_row_double_click(self, model_name):
        if self.current_step < 2: return
        self.last_selected_model = model_name
        self.populate_detail_view(model_name)
        self.tabview.set("상세")
        self.update_status_bar(f"'{model_name}'의 상세 정보를 표시합니다.")

    def get_cell_info(self, item_id, column_id):
        try:
            model_name = self.tree.item(item_id, "values")[0]
            current_value_str = self.tree.set(item_id, column_id)
            
            numbers = re.findall(r'(\d[\d,]*)', current_value_str)
            total_value = sum(int(n.replace(',', '')) for n in numbers)

            return self.get_cell_info_from_header(model_name, self.tree.heading(column_id, "text")) + (total_value,)
        except (ValueError, IndexError) as e:
            logging.warning(f"셀 정보 파싱 오류: {e} (선택한 헤더: {self.tree.heading(column_id, 'text')})")
            return None, None, None, None

    def get_cell_info_from_header(self, model_name, header_text):
        try:
            date_str, truck_str = header_text.split(' ')
            month, day = map(int, date_str.split('-'))
            truck_num = int(re.search(r'\d+', truck_str).group())
            year = self.processor.date_cols[0].year

            if self.processor.date_cols and self.processor.date_cols[0].month == 12 and month == 1:
                year += 1

            ship_date = datetime.date(year, month, day)
            return model_name, ship_date, truck_num
        except (ValueError, IndexError):
            return None, None, None

    def edit_shipment_value(self, item_id, column_id):
        model, date, truck_num, current_value = self.get_cell_info(item_id, column_id)
        if not all([model, date, truck_num is not None]): return

        dialog = ctk.CTkInputDialog(
            text=f"'{model}' {date.strftime('%m-%d')} {truck_num}차 출고량을 수동으로 수정/고정합니다.\n\n(현재값: {current_value:,.0f})",
            title="출고량 수동 고정"
        )
        new_value_str = dialog.get_input()

        if new_value_str is not None:
            try:
                new_value = int(new_value_str)
                if new_value < 0: raise ValueError
                self.update_fixed_shipment(model, date, truck_num, new_value)
                self.recalculate_with_fixed_values()
            except (ValueError, TypeError):
                messagebox.showerror("입력 오류", "유효한 숫자를 입력해주세요.", parent=self)

    def update_fixed_shipment(self, model, date, truck_num, qty):
        self.processor.fixed_shipments = [
            s for s in self.processor.fixed_shipments
            if not (s['model'] == model and s['date'] == date and s['truck_num'] == truck_num)
        ]
        if qty > 0:
            self.processor.fixed_shipments.append({'model': model, 'date': date, 'truck_num': truck_num, 'qty': qty})
        logging.info(f"고정 출고량 업데이트: {model}, {date}, {truck_num}차 -> {qty}개")

    def fix_shipment(self, model, date, truck_num, from_menu=False):
        col_name = f'출고_{truck_num}차_{date.strftime("%m%d")}'
        if self.processor.simulated_plan_df is not None and col_name in self.processor.simulated_plan_df.columns:
            current_value = self.processor.simulated_plan_df.loc[model, col_name]
            if current_value > 0:
                self.update_fixed_shipment(model, date, truck_num, int(current_value))
                if from_menu: self.recalculate_with_fixed_values()
            else:
                if from_menu: messagebox.showinfo("정보", "0인 값은 고정할 수 없습니다.", parent=self)

    def unfix_shipment(self, model, date, truck_num):
        self.update_fixed_shipment(model, date, truck_num, 0)
        self.recalculate_with_fixed_values()
        
    def _get_shortage_messages(self):
        df = self.processor.simulated_plan_df
        if df is None: return []
        
        inventory_cols = sorted([c for c in df.columns if isinstance(c, str) and c.startswith('재고_')])
        if not inventory_cols: return []

        shortage_messages = []
        sorted_models = self.processor.item_master_df.index
        for model in sorted_models:
            if model not in df.index: continue
            safety_stock = self.processor.item_master_df.loc[model, 'SafetyStock']
            for inv_col in inventory_cols:
                if df.loc[model, inv_col] < safety_stock:
                    date_str = inv_col.split('_')[1]
                    year = self.processor.date_cols[0].year
                    try: date_obj = datetime.datetime.strptime(f"{year}-{date_str}", "%Y-%m%d").date()
                    except ValueError: date_obj = datetime.datetime.strptime(f"{year+1}-{date_str}", "%Y-%m%d").date()
                    current_stock = df.loc[model, inv_col]
                    shortage_messages.append(f"{model}: {date_obj.strftime('%m/%d')} 재고({current_stock:,}) < 최소({safety_stock:,})")
                    break
        return shortage_messages
    
    def update_all_warnings_ui(self):
        unmet_logs = self.processor.unmet_demand_log
        shortage_messages = self._get_shortage_messages()
        log_count = len(unmet_logs)
        shortage_count = len(shortage_messages)
        total_warnings = log_count + shortage_count
        
        if total_warnings > 0:
            arrow = "▲ 접기" if self.warnings_visible else "▼ 펼치기"
            self.warnings_toggle_button.configure(text=f"🚨 시스템 경고 {total_warnings}건 ({arrow})")
            self.warnings_header_frame.grid()
        else:
            self.warnings_header_frame.grid_remove()
            self.warnings_container.grid_remove()
            return

        if self.warnings_visible:
            self.warnings_container.grid()
            
            if log_count > 0:
                self.unmet_demand_frame.grid()
                for widget in self.unmet_list_frame.winfo_children(): widget.destroy()
                for log in unmet_logs:
                    msg = f"{log['date'].strftime('%m/%d')} {log['model']}: {log['unmet_qty']:,}개 부족"
                    ctk.CTkLabel(self.unmet_list_frame, text=msg, font=self.font_small, anchor="w").pack(fill="x", padx=5)
            else:
                self.unmet_demand_frame.grid_remove()

            if shortage_count > 0:
                self.shortage_frame.grid()
                for widget in self.shortage_list_frame.winfo_children(): widget.destroy()
                for msg in shortage_messages:
                    ctk.CTkLabel(self.shortage_list_frame, text=msg, font=self.font_small, anchor="w").pack(fill="x", padx=5)
            else:
                self.shortage_frame.grid_remove()
        else:
            self.warnings_container.grid_remove()

    def populate_detail_view(self, model_name):
        for widget in self.detail_frame.winfo_children():
            widget.destroy()
        if self.processor.simulated_plan_df is None: return
        model_data = self.processor.simulated_plan_df.loc[model_name]

        fig, ax = plt.subplots(figsize=(12, 4))

        sim_start_date = self.processor.planning_start_date or self.processor.date_cols[0].date()
        dates = [d for d in self.processor.date_cols if d.date() >= sim_start_date]

        date_strs = [d.strftime('%m-%d') for d in dates]
        inventory = [model_data.get(f"재고_{d.strftime('%m%d')}", 0) for d in dates]
        production = [model_data.get(d, 0) for d in dates]
        shipment_cols = [c for c in model_data.index if isinstance(c, str) and c.startswith('출고_')]
        shipments_by_date = {d: 0 for d in dates}
        for col in shipment_cols:
            try:
                date_str_from_col = col[-4:]
                year_to_use = dates[0].year
                if dates[0].month == 12 and int(date_str_from_col[:2]) == 1:
                    year_to_use += 1

                date_obj = datetime.datetime.strptime(f"{year_to_use}-{date_str_from_col[:2]}-{date_str_from_col[2:]}", "%Y-%m-%d")

                matching_date = next((d for d in dates if d.date() == date_obj.date()), None)
                if matching_date:
                    shipments_by_date[matching_date] += model_data[col]
            except (ValueError, KeyError, IndexError): continue

        total_shipments = [shipments_by_date.get(d, 0) for d in dates]

        ax.plot(date_strs, inventory, marker='o', linestyle='-', label='예상 재고')
        ax.bar(date_strs, production, color='skyblue', label='생산량(수요)')
        ax.bar(date_strs, [-s for s in total_shipments], color='salmon', label='총출고량')
        safety_stock = self.processor.item_master_df.loc[model_name, 'SafetyStock']
        if safety_stock > 0:
            ax.axhline(y=safety_stock, color='r', linestyle='--', label=f'최소 재고 ({safety_stock:,})')
        ax.set_title(f"'{model_name}' 재고 및 입출고 추이", fontdict={'fontsize': 14})
        ax.set_xlabel("날짜"); ax.set_ylabel("수량")
        ax.legend(); ax.grid(True, which='both', linestyle='--', linewidth=0.5)
        plt.setp(ax.get_xticklabels(), rotation=45, ha="right"); fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=self.detail_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(fill='both', expand=True, padx=10, pady=10)

    def check_shipment_capacity(self):
        df, messages = self.processor.simulated_plan_df, []
        if df is None or not self.processor.date_cols: return

        all_shipment_cols = [col for col in df.columns if isinstance(col, str) and col.startswith('출고_')]
        grouped_cols = {}
        for col in all_shipment_cols:
            parts = col.split('_'); key = (parts[2], parts[1])
            if key not in grouped_cols: grouped_cols[key] = []
            grouped_cols[key].append(col)

        for (date_str, truck_num_str), cols in grouped_cols.items():
            try:
                year_to_use = self.processor.date_cols[0].year
                month_from_str = int(date_str[:2])
                if self.processor.date_cols[0].month == 12 and month_from_str == 1:
                    year_to_use += 1
                date_obj_key = datetime.datetime.strptime(f"{year_to_use}{date_str}", "%Y%m%d").date()

                default_pallets = self.config_manager.config.get('PALLETS_PER_TRUCK', 36)
                pallet_size = self.config_manager.config.get('PALLET_SIZE', 60)
                pallets_for_day = self.config_manager.config.get('DAILY_PALLET_OVERRIDES', {}).get(date_obj_key, default_pallets)

                total_shipped_pieces = df[cols].sum().sum()

                if total_shipped_pieces > 0 and pallet_size > 0:
                    total_pallets_loaded = 0
                    for item_code in df.index:
                        item_qty = df.loc[item_code, cols].sum()
                        if item_qty > 0:
                            total_pallets_loaded += math.ceil(item_qty / pallet_size)

                    if total_pallets_loaded > pallets_for_day:
                        date_obj_display = datetime.datetime.strptime(f"{datetime.date.today().year}{date_str}", "%Y%m%d")
                        messages.append(f"⚠️ {date_obj_display.strftime('%m-%d')} {truck_num_str}: 적재된 팔레트({total_pallets_loaded:,.0f}PL) > 트럭 한도({pallets_for_day:,.0f}PL)")
            except (ValueError, IndexError) as e:
                logging.warning(f"차량 용량 계산 중 오류 발생: {e}")
                continue

        if messages:
            messagebox.showwarning("출고 용량 초과 경고", "\n".join(messages))

    def update_status_bar(self, message="준비 완료"):
        self.status_bar.configure(text=f"현재 파일: {self.current_file} | 상태: {message}")
        logging.info(f"상태 업데이트: {message}")

    def load_settings_to_gui(self):
        for key, entry_widget in self.settings_entries.items():
            entry_widget.delete(0, 'end')
            entry_widget.insert(0, str(self.config_manager.config.get(key, '')))
        for i, cb in self.day_checkboxes.items():
            cb.select() if self.config_manager.config.get('DELIVERY_DAYS', {}).get(str(i), 'False') == 'True' else cb.deselect()

        default_path = os.path.join(os.path.expanduser('~'), 'Downloads')
        self.auto_save_path_var.set(self.config_manager.config.get('AUTO_SAVE_PATH', default_path))
        logging.info("UI에 설정값 로드 완료.")

    def change_auto_save_path(self):
        initial_dir = self.config_manager.config.get('AUTO_SAVE_PATH')
        new_path = filedialog.askdirectory(title="자동 저장 폴더를 선택하세요", initialdir=initial_dir)
        if new_path:
            self.config_manager.config['AUTO_SAVE_PATH'] = new_path
            self.config_manager.save_config()
            self.load_settings_to_gui()
            messagebox.showinfo("저장 경로 변경", f"자동 저장 경로가 다음으로 변경되었습니다:\n{new_path}", parent=self)

    def open_daily_truck_dialog(self):
        dialog = DailyTruckDialog(
            self,
            self.config_manager.config.get('DAILY_TRUCK_OVERRIDES', {}),
            self.config_manager.config.get('DAILY_PALLET_OVERRIDES', {}),
            self.config_manager.config.get('PALLETS_PER_TRUCK'),
            font_normal=self.font_normal
        )
        self.wait_window(dialog)
        if dialog.result is not None:
            self.config_manager.config['DAILY_TRUCK_OVERRIDES'] = dialog.result['trucks']
            self.config_manager.config['DAILY_PALLET_OVERRIDES'] = dialog.result['pallets']
            self.save_settings_and_recalculate()

    def open_holiday_dialog(self):
        current_holidays = [d for d in self.config_manager.config.get('NON_SHIPPING_DATES', []) if isinstance(d, datetime.date)]
        dialog = HolidayDialog(self, current_holidays, font_normal=self.font_normal)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.config_manager.config['NON_SHIPPING_DATES'] = dialog.result
            self.save_settings_and_recalculate()

    def open_safety_stock_dialog(self):
        self.processor._ensure_item_master_loaded()
        if self.processor.item_master_df is None: return
        dialog = SafetyStockDialog(self, self.processor.item_master_df, font_normal=self.font_normal, font_bold=self.font_bold)
        self.wait_window(dialog)
        if dialog.result is not None:
            self.processor.item_master_df = dialog.result
            self.processor.save_item_master()
            messagebox.showinfo("저장 완료", "최소 재고 설정이 저장되었습니다. '설정 저장 및 재계산'으로 계획에 반영하세요.")
            if self.current_step >=2:
                self.recalculate_with_fixed_values()

    def open_item_order_dialog(self):
        self.processor._ensure_item_master_loaded()
        if self.processor.item_master_df is None:
            messagebox.showwarning("오류", "먼저 생산계획을 불러와야 품목 정보를 설정할 수 있습니다.")
            return

        current_order = self.processor.item_master_df.index.tolist()
        highlight_set = set(self.processor.highlight_models)

        dialog = ItemOrderDialog(self, item_list=current_order, highlight_items=highlight_set, font_normal=self.font_normal)
        self.wait_window(dialog)

        if dialog.result:
            new_order = dialog.result
            self.processor.item_master_df = self.processor.item_master_df.reindex(new_order)
            self.processor.item_master_df['Priority'] = range(1, len(self.processor.item_master_df) + 1)
            self.processor.save_item_master()

            self.processor._load_item_master()

            messagebox.showinfo("저장 완료", "새로운 품목 순서가 Item.csv에 저장되었습니다.")

            self.filter_grid()

if __name__ == "__main__":
    try:
        config_manager = ConfigManager()
        app = ProductionPlannerApp(config_manager)
        app.mainloop()
    except Exception as e:
        logging.critical(f"Fatal error: {e}", exc_info=True)
        messagebox.showerror("치명적 오류", f"프로그램 실행에 실패했습니다.\n{e}")